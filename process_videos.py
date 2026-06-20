#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频下载、转码、文本识别、AI分析一体化流程脚本

用法:
  # 跑全量（两个视频 sheet），3 并发，失败重试 3 次
  python process_videos.py --concurrency 3 --retry 3

  # 指定 sheet 全量
  python process_videos.py --sheet "YouTube视频" --concurrency 2

  # 指定单条（extra.id 或 title）
  python process_videos.py --sheet "YouTube视频" --id 2143

  # 只跑某个阶段
  python process_videos.py --sheet "YouTube视频" --id 2143 --step download
  python process_videos.py --sheet "YouTube视频" --id 2143 --step transcode
  python process_videos.py --sheet "YouTube视频" --id 2143 --step transcribe
  python process_videos.py --sheet "YouTube视频" --id 2143 --step analyze

  # 强制重新下载
  python process_videos.py --sheet "YouTube视频" --id 2143 --force

  # 重跑上次报告中失败的任务
  python process_videos.py --retry-failed report_20260610_141800.json

  # 干跑（只列任务，不执行）
  python process_videos.py --dry-run

  # 偏移/限量（跳过前10个，只处理5个）
  python process_videos.py --offset 10 --limit 5
"""

from __future__ import annotations

import os
import sys
import re
import shutil
import importlib.util
import argparse
import subprocess
import logging
import json
import time
import traceback
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field, asdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

import colorama
colorama.init()

from dotenv import load_dotenv
import requests
import pandas as pd
from openpyxl import load_workbook

# 从 package.json 读取版本号（与 JS 版保持一致）
__version__ = "unknown"
try:
    _pkg = json.load(open(os.path.join(os.path.dirname(__file__), "package.json")))
    __version__ = _pkg.get("version", "unknown")
except Exception:
    pass

# ── 控制台单行动态显示 ──
from console_ui import (
    update_line, clear_line, text_bar, Spinner,
    parse_ytdlp_line, parse_ffmpeg_progress, reset_ffmpeg_state,
    fmt_size,
)

# --env-file 需在 load_dotenv 之前解析
_env_file = ".env"
if "--env-file" in sys.argv:
    _idx = sys.argv.index("--env-file")
    if _idx + 1 < len(sys.argv):
        _env_file = sys.argv[_idx + 1]
load_dotenv(dotenv_path=_env_file, override=True)

# ─────────────────────────────── 路径配置 ───────────────────────────────────

SCRIPT_DIR = Path(__file__).parent.resolve()  # 脚本自带资源（.env.example 等）
BASE_DIR = Path.cwd()                          # 用户数据路径（Excel、下载、输出等）

def _env_path(key: str, default: str) -> Path:
    """读取环境变量，如果是相对路径则相对于 BASE_DIR，绝对路径则直接使用"""
    val = os.getenv(key, default)
    p = Path(val)
    return p if p.is_absolute() else BASE_DIR / p

EXCEL_FILE = _env_path("EXCEL_FILE", "data/examples/website_split.xlsx")
COOKIES_DIR = _env_path("COOKIES_DIR", "data/cookies")

def _apply_output_dir(new_root: Path, *, log_func=None) -> None:
    """
    用 --output / OUTPUT_DIR 指定的根目录覆盖所有 7 个子目录常量。
    子目录名固定；调用后立即 mkdir。
    注意：必须用 global 声明，否则会创建局部变量导致其他模块看不到。
    """
    global OUTPUT_DIR, DOWNLOADS_DIR, TRANSCODED_DIR, TRANSCRIPTS_DIR
    global KEYWORDS_DIR, REPORTS_DIR, PROGRESS_DIR, LOGS_DIR
    OUTPUT_DIR = new_root
    DOWNLOADS_DIR   = OUTPUT_DIR / "downloads"
    TRANSCODED_DIR  = OUTPUT_DIR / "transcoded"
    TRANSCRIPTS_DIR = OUTPUT_DIR / "transcripts"
    KEYWORDS_DIR    = OUTPUT_DIR / "keywords"
    REPORTS_DIR     = OUTPUT_DIR / "reports"
    PROGRESS_DIR  = OUTPUT_DIR / "progress"
    LOGS_DIR        = OUTPUT_DIR / "logs"
    for _d in (DOWNLOADS_DIR, TRANSCODED_DIR, TRANSCRIPTS_DIR, KEYWORDS_DIR, REPORTS_DIR, PROGRESS_DIR, LOGS_DIR):
        _d.mkdir(parents=True, exist_ok=True)
    if log_func:
        log_func(f"输出根目录覆盖为: {OUTPUT_DIR}")


# ── 输出根目录 + 7 个固定子目录（子目录名不可通过 env 覆盖）──
OUTPUT_DIR = _env_path("OUTPUT_DIR", "output")
_apply_output_dir(OUTPUT_DIR)

YTDLP = os.getenv("YTDLP", "yt-dlp")
FFMPEG = os.getenv("FFMPEG", "ffmpeg")
FFPROBE = os.getenv("FFPROBE", "ffprobe")
# ── Whisper 共享参数（local、faster-whisper 和 service 通用） ──
WHISPER_BACKEND = os.getenv("WHISPER_BACKEND", "local")  # "service" / "faster-whisper" / "local"
WHISPER_TEMPERATURE = os.getenv("WHISPER_TEMPERATURE", "0.0")  # 推理温度 (0.0~1.0)
WHISPER_TEMPERATURE_INC = os.getenv("WHISPER_TEMPERATURE_INC", "0.2")  # 温度增量 (fallback 时升温步长)
WHISPER_OUTPUT_FORMAT = os.getenv("WHISPER_OUTPUT_FORMAT", "json")  # 输出格式: txt/vtt/srt/tsv/json/all (服务端映射到 response_format)

# ── Whisper 服务模式参数（独有） ──
WHISPER_SERVICE = os.getenv("WHISPER_SERVICE", "http://localhost:9588")  # 服务地址
# 服务端模型路径（ggml 文件，用于 POST /load 切换模型；留空则使用服务端当前加载的模型）
WHISPER_SERVICE_MODEL = os.getenv("WHISPER_SERVICE_MODEL", "")

# ── Whisper 本地模式参数（local / faster-whisper 通用） ──         //
WHISPER_TASK = os.getenv("WHISPER_TASK", "transcribe")  # 任务类型: transcribe/translate
WHISPER_MODEL = os.getenv("WHISPER_MODEL", "medium")  # 模型名: tiny/base/small/medium/large-v3/turbo
WHISPER_LANGUAGE = os.getenv("WHISPER_LANGUAGE", "")  # 语言: 设 zh 避免繁体混入; 留空=自动检测
WHISPER_MODEL_DIR = os.getenv("WHISPER_MODEL_DIR", "")  # 模型下载目录，留空=~/.cache/whisper（或 $XDG_CACHE_HOME/whisper）
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cpu")  # cpu / cuda
WHISPER_BEAM_SIZE = os.getenv("WHISPER_BEAM_SIZE", "5")  # beam 宽度 (温度=0 时生效, 越大越准)
WHISPER_BEST_OF = os.getenv("WHISPER_BEST_OF", "5")  # 候选数 (温度>0 时生效)
WHISPER_INITIAL_PROMPT: str = os.getenv("WHISPER_INITIAL_PROMPT", "")  # 初始提示词: 给首段音频提供词汇上下文, 提升专有名词识别; 示例见 .env.example
WHISPER_CONDITION_ON_PREV = os.getenv("WHISPER_CONDITION_ON_PREV", "False")  # 推荐 False: 每段独立解码, 避免长视频错误累积; True=前段文本传入(仅适合短音频)
WHISPER_FP16 = os.getenv("WHISPER_FP16", "False")  # CPU 应设为 False
WHISPER_THREADS = os.getenv("WHISPER_THREADS", "0")  # 线程数 (0=自动)

# ── faster-whisper 专用参数（backend=faster-whisper 时生效） ──
WHISPER_COMPUTE_TYPE = os.getenv("WHISPER_COMPUTE_TYPE", "int8")  # 计算精度: int8/float16/int8_float16/default/auto
WHISPER_VAD_FILTER = os.getenv("WHISPER_VAD_FILTER", "True")  # VAD 静音过滤 (True/False)
WHISPER_VAD_ONSET = os.getenv("WHISPER_VAD_ONSET", "0.5")  # VAD 灵敏度阈值 (0.0~1.0)
WHISPER_NUM_WORKERS = os.getenv("WHISPER_NUM_WORKERS", "1")  # CTranslate2 并行 worker 数

# ── FunASR 专用参数（backend=funasr 时生效） ──
# FunASR 专攻中文场景，WER ~5%（Whisper 中文 ~15%）。
# 需先安装: pip install funasr modelscope (cli)
#           或:   pip install funasr vllm fastapi uvicorn python-multipart (service, GPU 推荐)
FUNASR_MODE = os.getenv("FUNASR_MODE", "cli")  # "cli" = 本地 AutoModel; "service" = 远程 funasr-server (OpenAI 兼容 API)
FUNASR_MODEL = os.getenv("FUNASR_MODEL", "paraformer-zh")  # 主 ASR 模型: paraformer-zh / SenseVoiceSmall / Fun-ASR-Nano / Qwen3-ASR ...
FUNASR_VAD_MODEL = os.getenv("FUNASR_VAD_MODEL", "fsmn-vad")  # VAD 模型（留空=用主模型内置）
FUNASR_PUNC_MODEL = os.getenv("FUNASR_PUNC_MODEL", "ct-punc")  # 标点恢复（留空=不做）
FUNASR_SPK_MODEL = os.getenv("FUNASR_SPK_MODEL", "")  # 说话人分离（留空=不做）
FUNASR_EMOTION_MODEL = os.getenv("FUNASR_EMOTION_MODEL", "")  # 情感识别（留空=不做）
FUNASR_DEVICE = os.getenv("FUNASR_DEVICE", "cpu")  # cpu / cuda（GPU 强烈推荐）
FUNASR_QUANTIZE = os.getenv("FUNASR_QUANTIZE", "True")  # int8 量化（省 50% 内存, GPU 设 False）
FUNASR_BATCH_SIZE_S = os.getenv("FUNASR_BATCH_SIZE_S", "300")  # 动态批处理音频秒数 (60-600)
FUNASR_HOTWORD = os.getenv("FUNASR_HOTWORD", "")  # 热词（空格分隔, 显著提升专有名词）
FUNASR_LANGUAGE = os.getenv("FUNASR_LANGUAGE", "zh")  # 主语言（中文 zh, SenseVoice 配 auto 可自动检测 50+ 语种）
FUNASR_VAD_MAX_SEGMENT = os.getenv("FUNASR_VAD_MAX_SEGMENT", "20000")  # VAD 最大单段长度 (ms, 0=不切分)
FUNASR_SERVICE_URL = os.getenv("FUNASR_SERVICE_URL", "http://localhost:8899")  # funasr-server 地址
FUNASR_SERVICE_MODEL = os.getenv("FUNASR_SERVICE_MODEL", "iic/SenseVoiceSmall")  # 服务侧加载的模型 ID

_SERVICE_MODEL_LOADED: str | None = None  # 缓存的已加载模型，避免重复 /load
_FW_MODEL: object | None = None  # 缓存的 faster-whisper WhisperModel 实例
_FW_MODEL_CFG: str = ""  # 缓存的模型配置指纹 (model/device/compute_type 组合)

# ── CLI 覆盖占位（CLI 解析后由 apply_cli_overrides 填充）──
_cli_ai_prompt: str | None = None  # --ai-prompt
_resolved_whisper_extra_args: list[str] = []  # 解析后的参数数组
_resolved_funasr_extra_args: list[str] = []  # FunASR 额外参数（CLI > .env）


def _resolve_prompt_value(val: str | None) -> str:
    """解析提示词值：自动检测是文件路径还是内联文本。
    如果值对应的文件存在，读取文件内容；否则当文本直接返回。
    读取后自动将字面量 \\n \\t 转义为真正的换行/制表符。
    """
    if not val:
        return ""
    p = Path(val)
    if not p.is_absolute():
        p = BASE_DIR / p
    if p.is_file():
        content = p.read_text(encoding="utf-8").strip()
        return content.replace("\\n", "\n").replace("\\t", "\t")
    return val.replace("\\n", "\n").replace("\\t", "\t")


def _parse_extra_args(raw: str | None) -> list[str]:
    """解析 shell 字符串为参数数组。例如 '--beam_size 5' → ['--beam_size', '5']"""
    if not raw or not raw.strip():
        return []
    return raw.strip().split()


def _merge_whisper_args(base_args: list[str], extra_args: list[str]) -> list[str]:
    """合并 whisper 参数：基础参数 + 额外参数去重。
    额外参数优先级最高，如果基础参数中存在同名 key（--xxx），则移除基础参数中的该 key-value 对。
    """
    if not extra_args:
        return base_args
    extra_keys = {a for a in extra_args if a.startswith("--")}
    merged: list[str] = []
    i = 0
    while i < len(base_args):
        arg = base_args[i]
        if arg.startswith("--") and arg in extra_keys:
            # 跳过冲突的 key 及其 value
            if i + 1 < len(base_args) and not base_args[i + 1].startswith("-"):
                i += 2
            else:
                i += 1
            continue
        merged.append(arg)
        i += 1
    return merged + extra_args


def apply_cli_overrides(args: argparse.Namespace) -> None:
    """应用 CLI 覆盖：CLI > .env > 内置默认（在 parser.parse_args() 后调用）"""
    global WHISPER_INITIAL_PROMPT, _cli_ai_prompt, _resolved_whisper_extra_args, _resolved_funasr_extra_args

    # whisper-initial-prompt: CLI > .env > 内置默认
    if args.whisper_initial_prompt is not None:
        WHISPER_INITIAL_PROMPT = _resolve_prompt_value(args.whisper_initial_prompt)
    else:
        WHISPER_INITIAL_PROMPT = _resolve_prompt_value(WHISPER_INITIAL_PROMPT)

    # ai-prompt: CLI > .env > 内置默认（存入全局供 step_analyze 使用）
    if args.ai_prompt is not None:
        _cli_ai_prompt = _resolve_prompt_value(args.ai_prompt)

    # whisper-extra-args: CLI > .env
    raw_extra = args.whisper_extra_args or os.getenv("WHISPER_EXTRA_ARGS", "")
    _resolved_whisper_extra_args = _parse_extra_args(raw_extra)
    if _resolved_whisper_extra_args:
        with _print_lock:
            print(f"  whisper extra args: {' '.join(_resolved_whisper_extra_args)}", flush=True)

    # funasr-extra-args: CLI > .env
    raw_funasr_extra = getattr(args, "funasr_extra_args", None) or os.getenv("FUNASR_EXTRA_ARGS", "")
    _resolved_funasr_extra_args = _parse_extra_args(raw_funasr_extra)
    if _resolved_funasr_extra_args:
        with _print_lock:
            print(f"  funasr extra args: {' '.join(_resolved_funasr_extra_args)}", flush=True)

TRANSCODE_EXT = os.getenv("TRANSCODE_EXT", ".wav")
FFMPEG_TRANSCODE_ARGS = os.getenv("TRANSCODE_ARGS", "-vn -map_metadata -1 -map 0:a:0 -af loudnorm=I=-16:TP=-1.5:LRA=11:linear=true,aresample=resampler=soxr:osr=16000:osf=s16:dither_method=shibata -ac 1 -c:a pcm_s16le").split()

# ─────────────────────────────── Excel 字段映射 ─────────────────────────────

COL_ID = os.getenv("COL_ID", "extra.id")
COL_TITLE = os.getenv("COL_TITLE", "title")
COL_CONTENT = os.getenv("COL_CONTENT", "content")
COL_KEYWORDS = os.getenv("COL_KEYWORDS", "keywords")
COL_TENCENTVID = os.getenv("COL_TENCENTVID", "extra.tencentVid")
COL_BILIBILIBVID = os.getenv("COL_BILIBILIBVID", "extra.bilibiliBvid")
COL_YOUTUBEID = os.getenv("COL_YOUTUBEID", "extra.youtubeId")
COL_YOUKUID = os.getenv("COL_YOUKUID", "extra.youkuId")

# ──────────────────────────────── 平台配置 ──────────────────────────────────

# 平台 ID 列映射
_PLATFORM_COL_MAP = {
    "tencent":    COL_TENCENTVID,
    "tencentVid": COL_TENCENTVID,
    "bilibili":   COL_BILIBILIBVID,
    "bilibiliBvid": COL_BILIBILIBVID,
    "youtube":    COL_YOUTUBEID,
    "youtubeId":  COL_YOUTUBEID,
    "youku":      COL_YOUKUID,
    "youkuId":    COL_YOUKUID,
}

# 平台优先级
PLATFORM_PRIORITY = [p.strip() for p in os.getenv(
    "PLATFORM_PRIORITY", "bilibili,youtube,tencent,youku").split(",") if p.strip()]

# 视频 Sheet 列表
_VIDEO_SHEETS_RAW = os.getenv("VIDEO_SHEETS", "")
VIDEO_SHEETS = [s.strip() for s in _VIDEO_SHEETS_RAW.split(",") if s.strip()] if _VIDEO_SHEETS_RAW else []


# 平台 key → 环境变量前缀（使 .env 中的变量名简短可读）
_PKEY_ENV_PREFIX = {
    "tencent":    "TENCENT",
    "tencentVid": "TENCENT",
    "bilibili":   "BILIBILI",
    "bilibiliBvid": "BILIBILI",
    "youtube":    "YOUTUBE",
    "youtubeId":  "YOUTUBE",
    "youku":      "YOUKU",
    "youkuId":    "YOUKU",
}


def _build_platform_config() -> dict:
    """从环境变量动态构建平台配置字典"""
    config = {}
    for pkey in PLATFORM_PRIORITY:
        prefix = _PKEY_ENV_PREFIX.get(pkey, pkey.upper())
        cfg: dict = {
            "field": _PLATFORM_COL_MAP.get(pkey, f"extra.{pkey}"),
            "url_tpl": os.getenv(f"{prefix}_URL_TPL", ""),
        }

        # Cookie: cookies-from-browser 优先于 cookies file
        cfb = os.getenv(f"{prefix}_COOKIES_FROM_BROWSER", "")
        cookie_file = os.getenv(f"{prefix}_COOKIE_FILE", "")
        cfg["cookies_from_browser"] = cfb  # e.g. "chrome", "firefox", "edge"
        cfg["cookie_file"] = str(BASE_DIR / cookie_file) if cookie_file else None

        # Proxy（如 http://127.0.0.1:7890）
        proxy = os.getenv(f"{prefix}_PROXY", "")
        if proxy:
            cfg["proxy"] = proxy

        # Extra headers
        ua = os.getenv(f"{prefix}_USER_AGENT", "")
        extra_headers = []
        if ua:
            extra_headers += ["--user-agent", ua]
        if pkey == "bilibili":
            referer = os.getenv(f"{prefix}_REFERER", "")
            if referer:
                extra_headers += ["--add-header", f"Referer:{referer}"]
        if ua or (pkey == "bilibili" and os.getenv(f"{prefix}_REFERER", "")):
            extra_headers += ["--add-header", "Accept-Language:zh,en;q=0.9"]
        if extra_headers:
            cfg["extra_headers"] = extra_headers

        # Concurrent fragments
        cf = os.getenv(f"{prefix}_CONCURRENT_FRAGMENTS", "")
        if cf:
            cfg["concurrent_fragments"] = int(cf)

        # Extra args (YouTube-specific: JS runtime, remote components)
        if pkey == "youtube":
            js_rt = os.getenv(f"{prefix}_JS_RUNTIMES", "")
            rc = os.getenv(f"{prefix}_REMOTE_COMPONENTS", "")
            extra_args = []
            if js_rt:
                extra_args += ["--js-runtimes", js_rt]
            if rc:
                extra_args += ["--remote-components", rc]
            if extra_args:
                cfg["extra_args"] = extra_args

        # Format
        fmt = os.getenv(f"{prefix}_FORMAT", "")
        if fmt:
            cfg["format"] = fmt

        config[pkey] = cfg
    return config


PLATFORM_CONFIG = _build_platform_config()

# ─────────────────────────────── 日志配置 ───────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Excel 写锁（openpyxl 非线程安全）
_excel_lock = Lock()
# 控制台打印锁（并发时防止输出交错）
_print_lock = Lock()

# ─────────────────── 断点续跑 / 产物校验工具 ───────────────────
# transcribe 产物最小长度（字符），低于此视为残缺
MIN_TRANSCRIPT_CHARS = int(os.getenv("MIN_TRANSCRIPT_CHARS", "50"))
# analyze 产物最小长度（字符），低于此视为残缺
MIN_KEYWORDS_CHARS = int(os.getenv("MIN_KEYWORDS_CHARS", "5"))


def transcript_path(sheet_name: str, stem: str) -> Path:
    """识别文本落盘路径。"""
    d = TRANSCRIPTS_DIR / sheet_name
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{stem}.txt"


def keywords_path(sheet_name: str, stem: str) -> Path:
    """关键词落盘路径。"""
    d = KEYWORDS_DIR / sheet_name
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{stem}.txt"


def progress_path(sheet_name: str, stem: str) -> Path:
    """单任务 progress JSON 路径。"""
    d = PROGRESS_DIR / sheet_name
    d.mkdir(parents=True, exist_ok=True)
    return d / f"task_{stem}.json"


def safe_remove(path: Path | None) -> None:
    """安全删除文件（不存在/已删除不报错）。用于失败时清理残缺产物。"""
    if not path:
        return
    try:
        if path.exists():
            path.unlink()
    except OSError as e:
        log.warning(f"清理文件失败 {path}: {e}")


def validate_transcript_text(text: str | None) -> tuple[bool, str | None]:
    """transcribe 产物校验：非空 + 长度达标。返回 (是否有效, 错误信息)。"""
    if not text or not text.strip():
        return False, "识别文本为空"
    if len(text.strip()) < MIN_TRANSCRIPT_CHARS:
        return False, f"识别文本过短({len(text.strip())}<{MIN_TRANSCRIPT_CHARS})"
    return True, None


def validate_keywords_text(text: str | None) -> tuple[bool, str | None]:
    """analyze 产物校验：非空 + 长度达标。返回 (是否有效, 错误信息)。"""
    if not text or not text.strip():
        return False, "关键词为空"
    if len(text.strip()) < MIN_KEYWORDS_CHARS:
        return False, f"关键词过短({len(text.strip())}<{MIN_KEYWORDS_CHARS})"
    return True, None


def load_task_progress(sheet_name: str, stem: str) -> dict | None:
    """读取 progress JSON。文件不存在/解析失败返回 None。"""
    p = progress_path(sheet_name, stem)
    if not p.exists():
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        log.warning(f"progress JSON 解析失败 {p}: {e}")
        return None


def c(color: str, text: str) -> str:
    """返回带 ANSI 颜色码的文本（使用 colorama）"""
    colors = {
        "bold":    colorama.Style.BRIGHT,
        "dim":     colorama.Style.DIM,
        "yellow":  colorama.Fore.LIGHTYELLOW_EX,
        "cyan":    colorama.Fore.LIGHTCYAN_EX,
        "green":   colorama.Fore.LIGHTGREEN_EX,
        "red":     colorama.Fore.LIGHTRED_EX,
        "blue":    colorama.Fore.LIGHTBLUE_EX,
        "magenta": colorama.Fore.LIGHTMAGENTA_EX,
        "reset":   colorama.Style.RESET_ALL,
    }
    return colors.get(color, "") + text + colors["reset"]

def _print_long(msg: str | None, *, max_chars: int = 800, indent: str = "       ") -> None:
    """打印可能较长的错误信息，超长时截断并提示。

    当错误信息包含换行（如 stderr 多行输出）时，逐行缩进打印。
    """
    if not msg:
        return
    s = str(msg)
    if len(s) > max_chars:
        s = s[:max_chars] + f"...(truncated, total {len(msg)} chars)"
    for ln in s.splitlines():
        if ln.strip():
            print(c("red", f"{indent}{ln}"))
        else:
            print(indent)



# ─────────────────────────────── 总体进度 ───────────────────────────────────

@dataclass
class OverallProgress:
    """线程安全的总体进度追踪"""
    total: int
    completed: int = 0
    success: int = 0
    failed: int = 0
    partial: int = 0
    no_video: int = 0
    _lock: Lock = field(default_factory=Lock)

    def add_result(self, status: str):
        with self._lock:
            self.completed += 1
            if status == "success":
                self.success += 1
            elif status == "failed":
                self.failed += 1
            elif status == "partial":
                self.partial += 1
            elif status == "no_video":
                self.no_video += 1

    def summary_line(self) -> str:
        with self._lock:
            pct = self.completed / self.total * 100 if self.total else 0
            parts = [c("dim", f"[总进度 {self.completed}/{self.total} ({pct:.1f}%)]")]
            parts.append(c("green", f"✅{self.success}") if self.success > 0 else c("dim", "✅0"))
            parts.append(c("red", f"❌{self.failed}") if self.failed > 0 else c("dim", "❌0"))
            parts.append(c("yellow", f"⚠️{self.partial}") if self.partial > 0 else c("dim", "⚠️0"))
            parts.append(c("cyan", f"⏹️{self.no_video}") if self.no_video > 0 else c("dim", "⏹️0"))
            return "  ".join(parts)

    def position_label(self) -> str:
        """返回当前任务在总体中的序号标签"""
        with self._lock:
            return f"[{self.completed + 1}/{self.total}]"


# ─────────────────────────────── 格式化辅助 ─────────────────────────────────

def fmt_elapsed(seconds: float) -> str:
    """格式化耗时为可读字符串"""
    sec = round(seconds)
    if sec < 60:
        return f"{sec}s"
    m = sec // 60
    s = sec % 60
    return f"{m}m{s}s"


def fmt_duration(sec: float | None) -> str:
    """格式化视频时长 M:SS 或 H:MM:SS"""
    if not sec or sec <= 0:
        return ""
    m = int(sec // 60)
    s = int(sec % 60)
    if m < 60:
        return f"{m}:{s:02d}"
    h = m // 60
    rm = m % 60
    return f"{h}:{rm:02d}:{s:02d}"


# ─────────────────────────────── 数据结构 ───────────────────────────────────

@dataclass
class StepResult:
    status: str          # "success" | "failed" | "skipped"
    file: str | None = None
    error: str | None = None
    retries_used: int = 0


@dataclass
class TaskResult:
    sheet: str
    id_val: str          # extra.id 或 title 的字符串表示
    title: str
    platform: str | None
    video_url: str | None
    stem: str
    download: StepResult = field(default_factory=lambda: StepResult("skipped"))
    transcode: StepResult = field(default_factory=lambda: StepResult("skipped"))
    transcribe: StepResult = field(default_factory=lambda: StepResult("skipped"))
    analyze: StepResult = field(default_factory=lambda: StepResult("skipped"))
    overall_status: str = "pending"   # "success" | "partial" | "failed" | "no_video"
    error: str | None = None

    def to_dict(self) -> dict:
        d = asdict(self)
        return d


# ─────────────────────────────── 工具函数 ───────────────────────────────────

def safe_filename(name: str) -> str:
    safe = re.sub(r'[\\/:*?"<>|]', '_', str(name)).strip()
    # 防止路径遍历
    while '..' in safe:
        safe = safe.replace('..', '_')
    # 防止以 . 开头（Unix 隐藏文件）
    safe = re.sub(r'^\.+', '', safe)
    return safe or 'unknown'


def get_video_id(row: pd.Series) -> tuple[str | None, str | None]:
    for pkey in PLATFORM_PRIORITY:
        cfg = PLATFORM_CONFIG[pkey]
        val = row.get(cfg["field"])
        if pd.notna(val) and str(val).strip():
            return pkey, str(val).strip()
    return None, None


def build_url(pkey: str, vid: str) -> str:
    return PLATFORM_CONFIG[pkey]["url_tpl"].replace(f"{{{pkey}}}", vid)

# ═══════════════════════════════════════════════════════════════════
# URL 解析（--url 模式）— 将被注入到 process_videos.py
# ═══════════════════════════════════════════════════════════════════

_URL_PLATFORM_MAP = [
    {
        "platform": "bilibili",
        "pkey": "bilibili",
        "patterns": [
            re.compile(r"bilibili\.com/video/(BV[a-zA-Z0-9]{10})"),
            re.compile(r"b23\.tv/([a-zA-Z0-9]+)"),
            re.compile(r'player\.bilibili\.com/player\.html\?[^"\'\\s]*\\baid=(\\d+)'),
        ],
    },
    {
        "platform": "youtube",
        "pkey": "youtube",
        "patterns": [
            re.compile(
                r"(?:youtube\.com/(?:watch\?v=|embed/|shorts/|live/)|youtu\.be/)([a-zA-Z0-9_-]{11})"
            ),
        ],
    },
    {
        "platform": "tencent",
        "pkey": "tencent",
        "patterns": [
            re.compile(r"v\.qq\.com/x/page/([a-zA-Z0-9]+)\.html"),
            re.compile(r"v\.qq\.com/x/cover/[^/]+/([a-zA-Z0-9]+)\.html"),
            re.compile(r"[?&]vid=([a-zA-Z0-9]+)"),
        ],
    },
    {
        "platform": "youku",
        "pkey": "youku",
        "patterns": [
            re.compile(r"v\.youku\.com/v_show/id_([a-zA-Z0-9=]+)\.html"),
        ],
    },
]


def parse_url(url: str) -> dict | None:
    """
    解析视频 URL，返回 {"platform","pkey","video_id","watch_url"} 或 None。
    支持标准链接、短链接、内嵌 iframe URL。
    """
    if not url or not isinstance(url, str):
        return None
    url = url.strip()

    # 提取 iframe src
    iframe_match = re.search(r"""src=["']([^"']+)["']""", url)
    if iframe_match:
        url = iframe_match.group(1)

    for entry in _URL_PLATFORM_MAP:
        for pat in entry["patterns"]:
            m = pat.search(url)
            if m and m.group(1):
                return {
                    "platform": entry["platform"],
                    "pkey": entry["pkey"],
                    "video_id": m.group(1),
                    "watch_url": url,
                }
    return None



# ─────────────────────────────── 文件名去重 ──────────────────────────────────

# key: (sheet_name, pandas_row_index) → deduplicated stem
_STEM_CACHE: dict[tuple, str] = {}


def precompute_stems(df: pd.DataFrame, sheet_name: str) -> None:
    """为一个 sheet 的所有行预计算去重后的文件名。

    去重策略（同 sheet 内）：
    1. 优先用 id 作为 stem
    2. 同 sheet 内有重复 id → 用 id_title
    3. 仍重复 → 追加对应平台视频 ID（如 bvid）
    """
    from collections import Counter

    # Pass 1: base stems (id or title fallback)
    base_stems: dict[int, str] = {}
    for idx, row in df.iterrows():
        eid = row.get(COL_ID)
        if pd.notna(eid) and str(eid).strip():
            base_stems[idx] = safe_filename(str(int(float(eid))))
        else:
            base_stems[idx] = safe_filename(str(row.get(COL_TITLE, "unknown")))

    # Pass 2: resolve duplicates with id_title
    counts1 = Counter(base_stems.values())
    resolved: dict[int, str] = {}
    for idx, stem in base_stems.items():
        if counts1[stem] > 1:
            row = df.loc[idx]
            title_part = safe_filename(str(row.get(COL_TITLE, "")))
            resolved[idx] = f"{stem}_{title_part}" if title_part else stem
        else:
            resolved[idx] = stem

    # Pass 3: still duplicates? append platform video ID
    counts2 = Counter(resolved.values())
    for idx, stem in resolved.items():
        if counts2[stem] > 1:
            row = df.loc[idx]
            pkey, vid = get_video_id(row)
            if pkey and vid:
                resolved[idx] = f"{stem}_{safe_filename(vid)}"

    # Store
    for idx, stem in resolved.items():
        _STEM_CACHE[(sheet_name, idx)] = stem

def resolve_url_conflict(proposed_path):
    """URL 模式文件冲突处理：箭头键选择（覆盖/跳过/自定义）。"""
    import questionary as qy

    if not proposed_path.exists():
        return {"action": "proceed", "path": proposed_path}

    stem = proposed_path.stem
    parent = proposed_path.parent
    ext = proposed_path.suffix

    print(f"\n⚠️  文件已存在: {proposed_path}")

    action = qy.select(
        "如何处理?",
        choices=[
            {"name": "覆盖 (重新下载替换)", "value": "overwrite"},
            {"name": "跳过 (保留已有文件)", "value": "skip"},
            {"name": "自定义文件名", "value": "custom"},
        ],
    ).ask()

    if action == "skip":
        return {"action": "skip", "path": None}
    if action == "overwrite":
        return {"action": "proceed", "path": proposed_path}

    # custom name
    custom_name = qy.text("输入自定义文件名 (不含扩展名):", default=stem).ask()
    if not custom_name or not custom_name.strip():
        print("文件名不能为空，使用默认名称")
        return {"action": "proceed", "path": proposed_path}
    new_path = parent / f"{safe_filename(custom_name)}{ext}"
    return resolve_url_conflict(new_path)



def stem_name(row: pd.Series, sheet_name: str = "") -> str:
    """获取去重后的文件名 stem。优先使用预计算缓存。"""
    idx = row.name  # pandas row index
    if sheet_name:
        cache_key = (sheet_name, idx)
        if cache_key in _STEM_CACHE:
            return _STEM_CACHE[cache_key]
    # Fallback: 无缓存时用 id 或 title
    eid = row.get(COL_ID)
    if pd.notna(eid) and str(eid).strip():
        return safe_filename(str(int(float(eid))))
    return safe_filename(str(row.get(COL_TITLE, "unknown")))


def row_key(row: pd.Series) -> str:
    """返回行的唯一键 (COL_ID 或 COL_TITLE)"""
    eid = row.get(COL_ID)
    if pd.notna(eid) and str(eid).strip():
        return str(int(float(eid)))
    return str(row.get(COL_TITLE, "unknown"))


def find_downloaded_file(dl_dir: Path, stem: str) -> Path | None:
    for p in dl_dir.iterdir():
        if p.stem == stem:
            return p
    return None


# ─────────────────────────────── 重试机制 ───────────────────────────────────

RETRYABLE_ERRORS = (
    subprocess.TimeoutExpired,
    requests.exceptions.Timeout,
    requests.exceptions.ConnectionError,
    requests.exceptions.HTTPError,
    ConnectionError,
    TimeoutError,
    OSError,            # 网络层面的错误
)

NON_RETRYABLE_ERRORS = (
    FileNotFoundError,
    PermissionError,
    ValueError,
    KeyError,
    TypeError,
)


def is_retryable(error: Exception) -> bool:
    """判断异常是否可重试"""
    if isinstance(error, NON_RETRYABLE_ERRORS):
        return False
    if isinstance(error, RETRYABLE_ERRORS):
        return True
    # subprocess.CalledProcessError：检查 stderr 是否包含永久错误
    if isinstance(error, subprocess.CalledProcessError):
        stderr_lower = (error.stderr or "").lower()
        non_retry_keywords = {
            "404", "403", "401", "unavailable", "private video",
            "video is not available", "this video is no longer",
            "removed", "deleted", "invalid url", "unsupported url",
        }
        for kw in non_retry_keywords:
            if kw in stderr_lower:
                return False
        return True
    return False  # 未知异常默认不重试


def retry_call(fn, *args, max_retries: int = 3, base_delay: float = 5.0,
               task_label: str = "", **kwargs) -> tuple[any, int, str | None]:
    """
    带重试的函数调用。
    返回 (返回值, 已用重试次数, 错误信息)。
    成功时 error 为 None；全部重试失败后抛出最后一个异常。
    """
    last_error = None
    for attempt in range(max_retries + 1):
        try:
            result = fn(*args, **kwargs)
            return result, attempt, None
        except Exception as e:
            last_error = str(e)
            if not is_retryable(e):
                raise
            if attempt < max_retries:
                delay = base_delay * (2 ** attempt)  # 5, 10, 20...
                log.warning(
                    f"[{task_label}] 第 {attempt + 1}/{max_retries + 1} 次尝试失败，{delay:.0f}s 后重试: {e}"
                )
                time.sleep(delay)
            else:
                raise
    # 理论上不会到这里
    raise RuntimeError(last_error or "unknown")


# ─────────────────────────────── 进度显示 ───────────────────────────────────

def run_with_progress(cmd: list[str], label: str, parser_fn, timeout: int = 600, extra_env: dict[str, str] | None = None):
    """
    执行命令并实时输出进度（单行动态刷新）。
    parser_fn 接收每一行，返回进度字符串（使用 ANSI escape 单行刷新）。
    返回 (完整输出, returncode)。
    """
    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    if extra_env:
        env.update(extra_env)
    proc = subprocess.Popen(
        cmd, stdin=subprocess.DEVNULL,
        stderr=subprocess.STDOUT, stdout=subprocess.PIPE,
        text=True, encoding="utf-8", errors="replace",
        env=env,
    )

    output_lines: list[str] = []
    start = time.monotonic()

    for line in proc.stdout:
        output_lines.append(line)
        progress = parser_fn(line)
        if progress:
            # 单行动态刷新，不换行
            update_line(f"  [{label}] {progress}")

        # 同步超时检测：每读一行检查一次
        if timeout > 0 and time.monotonic() - start > timeout:
            proc.kill()
            proc.wait()
            clear_line()
            sys.stderr.write('\n')
            raise subprocess.TimeoutExpired(cmd, timeout, output="".join(output_lines))

    proc.wait()
    # 进度完成，清除当前行并换行
    clear_line()
    sys.stderr.write('\n')
    return "".join(output_lines), proc.returncode


def parse_ytdlp_progress(line: str) -> str | None:
    """解析 yt-dlp 进度行，返回单行进度字符串（含进度条）。
    使用 console_ui.parse_ytdlp_line 做结构化解析，再用 text_bar 格式化。
    """
    parsed = parse_ytdlp_line(line)
    if not parsed:
        return None
    if parsed['type'] == 'dest':
        return f"↓ 下载 {parsed['ext']}..."
    if parsed['type'] == 'merge':
        return "↔ 合并中..."
    if parsed['type'] == 'progress':
        pct = parsed['percent']
        bar = text_bar(pct, 18)
        size_str = f"{parsed['downloaded']}{parsed['downloaded_unit']}"
        speed_str = f"{parsed['speed']}{parsed['speed_unit']}/s" if parsed['speed'] > 0 else '---'
        eta_str = parsed['eta'] or '--:--'
        return f"{bar} {pct}% | {size_str} @ {speed_str} | ETA {eta_str}"
    return None


def get_duration(filepath: Path) -> float | None:
    """用 ffprobe 获取媒体时长（秒）。失败返回 None。"""
    try:
        result = subprocess.run(
            [FFPROBE, "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(filepath)],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30,
        )
        duration_str = result.stdout.strip()
        if duration_str:
            return float(duration_str)
    except Exception:
        pass
    return None


def make_ffmpeg_parser(total_duration: float | None) -> callable:
    """根据总时长创建 ffmpeg 进度解析器（使用 -progress pipe:1 格式）"""
    def _parse(line: str) -> str | None:
        prog = parse_ffmpeg_progress(line, total_duration or 0)
        if not prog:
            return None
        bar = text_bar(prog['percent'], 18)
        elapsed_str = f"{int(prog['elapsed'] // 60):02d}:{int(prog['elapsed'] % 60):02d}"
        parts = [f"{bar} {prog['percent']}% | {elapsed_str}"]
        if prog.get('total_size', 0) > 0:
            parts.append(fmt_size(prog['total_size']))
        if prog.get('speed', 0) > 0:
            parts.append(f"{prog['speed']:.1f}x")
        return " ".join(parts)
    return _parse


def step_analyze(
    text: str,
    max_retries: int, retry_delay: float,
    timeout: int = 300,
    label: str = "analyze",
) -> tuple[str | None, int, str | None]:
    """调用 OpenAI 兼容 API 对识别文本做关键词归纳。
    返回 (keywords_text, retries_used, error_msg)。
    """
    if not text or not text.strip():
        return None, 0, "识别文本为空，跳过 AI 分析"

    ai_start = time.time()
    api_key = os.getenv("AI_API_KEY", "")
    base_url = os.getenv("AI_BASE_URL", "")
    model = os.getenv("AI_MODEL", "")
    # CLI > .env > 内置默认；_resolve_prompt_value 自动处理文件路径和 \n 转义
    prompt_tpl = (
        _cli_ai_prompt
        or _resolve_prompt_value(os.getenv("AI_PROMPT_TPL"))
        or """你是生物医药多语言内容分析与ASR纠错专家。对以下视频转录文本提取搜索关键词。

【第一步：语义修正与术语消歧】语音识别（Whisper）在处理专业内容时极易出错。请在理解上下文的基础上，激活生物医药专业词典进行以下修正（仅内部推理使用，不改变原文语义，不添加新内容）：
- 同音/近音错字：如"冻存"误为"洞存"、"储存"误为"铸存"、"传代"误为"传带"、"复苏"误为"复舒"、"抗体"误为"康体"、"细胞株"误为"细胞珠"、"培养基"误为"培养鸡"、"质粒"误为"智力"、"表达量"误为"表大量"。
- 形近字混淆：如"印迹"误为"印记"、"缓冲液"误为"缓冲夜"、"核酸"误为"核算"、"测序"误为"侧序"。
- 英文缩写与发音误判：如将"PCR"误识为中文或乱码，将"CRISPR"误识为"克里斯普"，需根据上下文还原为标准英文缩写。
边界约束：仅修正确实存在明显错误的词汇，保持原文行文逻辑不变。

【第二步：关键词提取规则】基于修正后的文本，严格遵循以下语言判定与提取规则：

【语言判定】
- 统计文本中的中文字符数与英文字母数
- 中文占比 > 60% → 按纯中文规则处理
- 英文占比 > 60% → 按纯英文规则处理
- 两者均不满足 → 按中英混合规则处理

【纯中文内容】
- 只提取文本中明确出现或直接体现的、具备实际搜索价值的关键词，数量适中，涵盖提供的内容即可，但是不能重复和凭空捏造、联想、扩展，用英文逗号分隔
- 关键词必须全部是中文，绝对不能翻译成英文
- 优先提取 2-8 字的专有名词、技术短语或核心概念
- 过滤单字及无意义泛词（如"的""是""这个""一个"等）

【纯英文内容】
- 只提取文本中明确出现或直接体现的、具备实际搜索价值的关键词，数量适中，涵盖提供的内容即可，但是不能重复和凭空捏造、联想、扩展，用英文逗号分隔
- 关键词必须全部是英文，绝对不能翻译成中文
- 优先提取 2-8 词的专业术语、基因/蛋白名称或实验方法等
- 过滤单字及无意义泛词（如"the""this""is""a"等）

【中英混合内容】
- 只提取文本中明确出现或直接体现的、具备实际搜索价值的关键词，数量适中，涵盖提供的内容即可，但是不能重复和凭空捏造、联想、扩展
- 语种隔离原则：中文关键词必须是中文，英文关键词必须是英文，互不翻译且严禁中英混杂
- 排序原则：中文关键词置于前，英文关键词置于后，统一用英文逗号分隔

通用规则：严禁凭空联想、扩展或编造文本中未出现的概念，宁少勿多。最终只输出以英文逗号分隔的关键词列表，不包含任何解释性文字。这是内容：{content}"""
    )
    ai_timeout = timeout

    if not api_key or not base_url or not model:
        return None, 0, "AI 配置不完整（缺少 AI_API_KEY / AI_BASE_URL / AI_MODEL）"

    # 组装提示词（Python str.replace 不存在 JS 的 $ 特殊字符问题）
    prompt = prompt_tpl.replace("{content}", text)

    # AI_DEBUG=true 时打印实际发送的 prompt 和返回内容，排查关键词质量问题
    ai_debug = os.getenv("AI_DEBUG", "").lower() == "true"
    if ai_debug:
        with _print_lock:
            print(f"  [{label}] 🔍 AI_DEBUG prompt({len(prompt)} chars): {prompt[:500]}...", flush=True)
            print(f"  [{label}] 🔍 AI_DEBUG transcript({len(text)} chars): {text[:200]}...", flush=True)

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": float(os.getenv("AI_TEMPERATURE", "0.3")),
    }

    import urllib.request
    import urllib.parse
    import json as _json

    url = base_url.rstrip("/") + "/chat/completions"
    with _print_lock:
        print(f"  [{label}] AI 请求 URL: {url}", flush=True)

    last_err = None
    spinner = Spinner()
    spinner.start(f"[{label}] AI 分析中")
    try:
        for attempt in range(1, max_retries + 2):  # 首次 + max_retries 次重试
            try:
                # 每次尝试重新创建 request（避免复用已消费的 data 流）
                data = _json.dumps(payload).encode("utf-8")
                req = urllib.request.Request(url, data=data, headers=headers, method="POST")
                with urllib.request.urlopen(req, timeout=ai_timeout if ai_timeout > 0 else None) as resp:
                    resp_body = resp.read().decode("utf-8")
                    status = resp.status
                if status != 200:
                    raise Exception(f"HTTP {status}: {resp_body[:300]}")
                body = _json.loads(resp_body)
                content = (body.get("choices") or [{}])[0].get("message", {}).get("content", "")
                result = content.strip(), attempt - 1, None
                spinner.stop()
                if ai_debug and result[0]:
                    with _print_lock:
                        print(f"  [{label}] 🔍 AI_DEBUG response({len(result[0])} chars): {result[0][:500]}", flush=True)
                with _print_lock:
                    elapsed = time.monotonic() - ai_start
                    m, s = divmod(int(elapsed), 60)
                    elapsed_str = f"{m}m{s:02d}s" if m > 0 else f"{s}s"
                    print(f"  [{label}] {c('green', 'AI 分析完成')} ({elapsed_str}, {len(result[0])} 字符)", flush=True)
                return result
            except Exception as e:
                err_str = str(e)[:500]
                last_err = err_str
                if attempt <= max_retries:  # 还有重试机会
                    sleep_sec = retry_delay * (2 ** (attempt - 1))
                    with _print_lock:
                        print(f"  [{label}] AI 第 {attempt} 次{c('red', '失败')}：{err_str[:100]}，{sleep_sec:.0f}s 后重试...", flush=True)
                    # 重试前重启 spinner
                    spinner.stop()
                    time.sleep(min(sleep_sec, 30))
                    spinner.start(f"[{label}] AI 分析中")
                else:
                    break
    finally:
        spinner.stop()

    return None, max_retries + 1, f"AI 分析失败（重试 {max_retries+1} 次）：{last_err} [url: {url}]"



def _cleanup_partial_files(dl_dir: Path, stem: str) -> None:
    """清理下载残留文件（.part 和 .ytdl）"""
    for pattern in [f"{stem}.*.part", f"{stem}.*.ytdl", f"{stem}.part", f"{stem}.ytdl"]:
        for f in dl_dir.glob(pattern):
            try:
                f.unlink()
                with _print_lock:
                    print(f"  [{stem}] 已清理残留文件: {f.name}", flush=True)
            except OSError:
                pass


def step_download(
    row: pd.Series, sheet_name: str,
    max_retries: int, retry_delay: float, force: bool,
    timeout: int = 1800,
) -> tuple[Path | None, int, str | None]:
    """下载视频。返回 (文件路径, 重试次数, 错误信息)"""
    pkey, vid = get_video_id(row)
    stem = stem_name(row, sheet_name)

    if not pkey:
        return None, 0, "无可用视频 ID"

    dl_dir = DOWNLOADS_DIR / sheet_name
    dl_dir.mkdir(parents=True, exist_ok=True)

    if not force:
        existing = find_downloaded_file(dl_dir, stem)
        if existing:
            # 清理该 stem 的残留 .part/.ytdl（如果有过中断下载）
            _cleanup_partial_files(dl_dir, stem)
            with _print_lock:
                print(c("dim", f"  [{stem}] 已存在 {existing.name}，跳过下载"), flush=True)
            return existing, 0, None

    url = build_url(pkey, vid)
    # 开始前清理残留的 .part/.ytdl（上次中断的下载）
    _cleanup_partial_files(dl_dir, stem)

    with _print_lock:
        print(f"  [{stem}] {c('cyan', '开始下载')} (平台={pkey})", flush=True)
        print(f"  [{stem}] {url}", flush=True)

    download_start = time.monotonic()

    # -X utf8: 强制 Python UTF-8 模式，修复 Windows 下 requests 库 latin-1 编码错误
    base_cmd = [sys.executable, "-X", "utf8", "-m", "yt_dlp"] if YTDLP == "yt-dlp" else [YTDLP]
    cmd = base_cmd + [
        "--no-update",             # 抑制版本过期警告（避免 Windows latin-1 编码报错）
        "--socket-timeout", "60",  # 增大 socket 超时（默认 20s，腾讯视频等站点易超时）
        url,
        "-o", str(dl_dir / f"{stem}.%(ext)s"),
        "--no-playlist",
        "--newline",               # 进度行以 \n 结尾，确保逐行可读
        "--merge-output-format", "mp4",
        "-f", PLATFORM_CONFIG[pkey].get("format", "bestvideo+bestaudio/best"),
    ]

    # 强制覆盖已有文件（--force 时 yt-dlp 也需忽略本地文件）
    if force:
        cmd += ["--force-overwrites"]

    # 平台级并发分片（加速下载）
    cf = PLATFORM_CONFIG[pkey].get("concurrent_fragments")
    if cf:
        cmd += ["--concurrent-fragments", str(cf)]

    # Cookie: cookies-from-browser 优先于 cookies file
    cfb = PLATFORM_CONFIG[pkey].get("cookies_from_browser", "")
    cookie_file = PLATFORM_CONFIG[pkey].get("cookie_file")
    if cfb:
        cmd += ["--cookies-from-browser", cfb]
    elif cookie_file and Path(cookie_file).exists():
        cmd += ["--cookies", cookie_file]

    # User-Agent（可选，解决腾讯视频等站点编码问题）
    user_agent = PLATFORM_CONFIG[pkey].get("user_agent", "")
    if user_agent:
        cmd += ["--user-agent", user_agent]

    # Proxy（如 http://127.0.0.1:7890）
    # 同时设置子进程环境变量，确保 yt-dlp 内的 node/ejs 等子进程也走代理
    proxy = PLATFORM_CONFIG[pkey].get("proxy", "")
    extra_env: dict[str, str] = {}
    if proxy:
        cmd += ["--proxy", proxy]
        extra_env["HTTPS_PROXY"] = proxy
        extra_env["HTTP_PROXY"] = proxy

    extra_headers = PLATFORM_CONFIG[pkey].get("extra_headers", [])
    if extra_headers:
        cmd += extra_headers

    extra_args = PLATFORM_CONFIG[pkey].get("extra_args", [])
    if extra_args:
        cmd += extra_args

    def _run():
        stderr_text, rc = run_with_progress(cmd, stem, parse_ytdlp_progress, timeout=timeout, extra_env=extra_env)
        if rc != 0:
            raise subprocess.CalledProcessError(rc, cmd, stderr=stderr_text)

    try:
        _run, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            _cleanup_partial_files(dl_dir, stem)
            return None, retries_used, err
    except Exception as e:
        stderr_text = ""
        if isinstance(e, subprocess.CalledProcessError):
            stderr_text = (e.stderr or "")[-2000:]
        log.error(f"[{stem}] yt-dlp 下载失败:\n{stderr_text or str(e)}")
        retries = max_retries if isinstance(e, subprocess.CalledProcessError) else 0
        _cleanup_partial_files(dl_dir, stem)
        return None, retries, (stderr_text or str(e))[:500]

    downloaded = find_downloaded_file(dl_dir, stem)
    if downloaded:
        elapsed = time.monotonic() - download_start
        size_mb = downloaded.stat().st_size / 1024 / 1024
        dur = get_duration(downloaded)
        dur_str = f", {fmt_duration(dur)}" if dur else ""
        with _print_lock:
            print(f"  [{stem}] {c('green', '下载完成')} -> {downloaded.name} ({size_mb:.1f} MB, {fmt_elapsed(elapsed)}{dur_str})", flush=True)
    else:
        log.error(f"[{stem}] 下载后找不到文件")
        return None, retries_used, "下载后找不到文件"
    return downloaded, 0, None


def step_transcode(
    src_file: Path, sheet_name: str,
    max_retries: int, retry_delay: float, force: bool,
    timeout: int = 1200,
    out_stem: str | None = None,
) -> tuple[Path | None, int, str | None]:
    """转码。返回 (转码文件路径, 重试次数, 错误信息)"""
    tc_dir = TRANSCODED_DIR / sheet_name
    tc_dir.mkdir(parents=True, exist_ok=True)
    stem = out_stem if out_stem else src_file.stem
    out_file = tc_dir / (stem + TRANSCODE_EXT)
    stem = src_file.stem

    if not force and out_file.exists() and out_file.stat().st_size > 0:
        # 如果源文件比转码文件更新（重新下载过），强制重转码
        if src_file.stat().st_mtime > out_file.stat().st_mtime:
            with _print_lock:
                print(c("yellow", f"  [{stem}] 源文件已更新（下载时间晚于转码时间），重新转码"), flush=True)
        else:
            with _print_lock:
                print(c("dim", f"  [{stem}] 已存在转码文件，跳过"), flush=True)
            return out_file, 0, None

    with _print_lock:
        print(f"  [{stem}] {c('blue', '开始转码')} -> {out_file.name}", flush=True)

    # 获取源文件时长用于百分比计算
    total_dur = get_duration(src_file)
    transcode_start = time.monotonic()
    reset_ffmpeg_state()

    def _run():
        parser = make_ffmpeg_parser(total_dur)
        cmd = [FFMPEG, "-y", "-i", str(src_file), "-progress", "pipe:1", "-nostats"] + FFMPEG_TRANSCODE_ARGS + [str(out_file)]
        stderr_text, rc = run_with_progress(cmd, stem, parser, timeout=timeout)
        if rc != 0:
            raise subprocess.CalledProcessError(rc, cmd, stderr=stderr_text)
        return out_file

    try:
        result, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        elapsed = time.monotonic() - transcode_start
        size_mb = out_file.stat().st_size / 1024 / 1024
        dur_str = f", {fmt_duration(total_dur)}" if total_dur else ""
        with _print_lock:
            print(f"  [{stem}] {c('green', '转码完成')} -> {out_file.name} ({size_mb:.1f} MB, {fmt_elapsed(elapsed)}{dur_str})", flush=True)
        return result, 0, None
    except Exception as e:
        stderr_text = ""
        if isinstance(e, subprocess.CalledProcessError):
            stderr_text = (e.stderr or "")[-2000:]
        log.error(f"[{stem}] ffmpeg 转码失败:\n{stderr_text or str(e)}")
        return None, max_retries, (stderr_text or str(e))[:500]


def _check_whisper_available() -> bool:
    """检测 whisper 是否可用（按 WHISPER_BACKEND 判断）。

    注意：whisper / whisper-ctranslate2 / funasr 等 Python CLI 的 --help 或 import
    均会触发框架初始化（Hydra / CTranslate2 / 模型注册），耗时 5-30 秒，
    不适合用作预检。改用轻量检测：
      - CLI 后端 → shutil.which() 检查可执行文件是否存在（< 0.3 秒）
      - Python API 后端 → importlib.util.find_spec() 检查包是否已安装（< 0.3 秒）
    """
    if WHISPER_BACKEND == "local":
        if shutil.which("whisper"):
            return True
        log.error("本地 whisper CLI 不可用，请确认: pip install openai-whisper")
        return False
    elif WHISPER_BACKEND == "faster-whisper":
        if importlib.util.find_spec("faster_whisper") is not None:
            return True
        log.error("faster-whisper 不可用，请确认: pip install faster-whisper")
        return False
    elif WHISPER_BACKEND == "funasr":
        if FUNASR_MODE == "service":
            try:
                requests.get(FUNASR_SERVICE_URL, timeout=3)
                return True
            except Exception:
                log.error(f"funasr-server 不可用（{FUNASR_SERVICE_URL}），请确认服务已启动: funasr-server")
                return False
        if importlib.util.find_spec("funasr") is not None:
            return True
        log.error("funasr 不可用，请确认: pip install funasr modelscope")
        return False
    else:
        try:
            r = requests.get(WHISPER_SERVICE, timeout=3)
            return True
        except Exception:
            return False


def check_environment(steps: list[str]) -> dict:
    """检测本次执行涉及的工具/服务是否可用。

    返回 {"all_ok": bool, "issues": [str], "ytdlp": bool, ...}"""
    result: dict = {
        "ytdlp": True, "ffmpeg": True, "ffprobe": True,
        "whisper": True, "ai": True, "all_ok": True, "issues": [],
    }

    if "download" in steps:
        if shutil.which(YTDLP) is None:
            result["ytdlp"] = False
            result["all_ok"] = False
            result["issues"].append(f"yt-dlp 不可用 ({YTDLP})")

    if "transcode" in steps:
        if shutil.which(FFMPEG) is None:
            result["ffmpeg"] = False
            result["all_ok"] = False
            result["issues"].append(f"ffmpeg 不可用 ({FFMPEG})")

    if "transcribe" in steps:
        if not _check_whisper_available():
            result["whisper"] = False
            result["all_ok"] = False
            if WHISPER_BACKEND == "local":
                backend = "本地 whisper CLI"
            elif WHISPER_BACKEND == "faster-whisper":
                backend = "faster-whisper (faster_whisper)"
            elif WHISPER_BACKEND == "funasr":
                if FUNASR_MODE == "service":
                    backend = f"funasr/service ({FUNASR_SERVICE_URL})"
                else:
                    backend = f"funasr/cli ({FUNASR_MODEL})"
            else:
                backend = f"service {WHISPER_SERVICE}"
            result["issues"].append(f"whisper 不可用 ({backend})")

    if "analyze" in steps:
        ai_enabled = os.getenv("AI_ENABLED", "true").lower() == "true"
        ai_key = os.getenv("AI_API_KEY", "")
        ai_url = os.getenv("AI_BASE_URL", "")
        if not ai_enabled:
            result["ai"] = False
            result["all_ok"] = False
            result["issues"].append("AI 分析已禁用 (AI_ENABLED=false)")
        elif not ai_key or not ai_url:
            result["ai"] = False
            result["all_ok"] = False
            result["issues"].append("AI 分析配置不完整（缺少 AI_API_KEY / AI_BASE_URL）")

    return result


def step_transcribe(
    audio_file: Path,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """调用 whisper 识别（支持 local / faster-whisper / service 三种后端）。返回 (文本, 重试次数, 错误信息)"""
    stem = audio_file.stem

    if not _check_whisper_available():
        if WHISPER_BACKEND == "local":
            backend_info = "本地 whisper CLI"
        elif WHISPER_BACKEND == "faster-whisper":
            backend_info = "faster-whisper (faster_whisper)"
        elif WHISPER_BACKEND == "funasr":
            if FUNASR_MODE == "service":
                backend_info = f"funasr/service ({FUNASR_SERVICE_URL})"
            else:
                backend_info = f"funasr/cli ({FUNASR_MODEL})"
        else:
            backend_info = f"service {WHISPER_SERVICE}"
        msg = f"whisper 不可用 ({backend_info})"
        log.warning(f"[{stem}] {msg}")
        return None, 0, msg

    file_size_mb = audio_file.stat().st_size / (1024 * 1024)
    if WHISPER_BACKEND == "local":
        mode_label = "local"
        model_label = WHISPER_MODEL
        lang_label = WHISPER_LANGUAGE if WHISPER_LANGUAGE else "auto"
    elif WHISPER_BACKEND == "faster-whisper":
        mode_label = "faster-whisper"
        model_label = f"{WHISPER_MODEL}/{WHISPER_COMPUTE_TYPE}"
        lang_label = WHISPER_LANGUAGE if WHISPER_LANGUAGE else "auto"
    elif WHISPER_BACKEND == "funasr":
        mode_label = f"funasr/{FUNASR_MODE}"
        if FUNASR_MODE == "service":
            model_label = Path(FUNASR_SERVICE_MODEL).name if FUNASR_SERVICE_MODEL else "(server default)"
        else:
            _m = [FUNASR_MODEL]
            if FUNASR_VAD_MODEL:    _m.append(FUNASR_VAD_MODEL)
            if FUNASR_PUNC_MODEL:   _m.append(FUNASR_PUNC_MODEL)
            model_label = "+".join(_m)
        lang_label = FUNASR_LANGUAGE if FUNASR_LANGUAGE else "auto"
    else:
        mode_label = "service"
        model_label = Path(WHISPER_SERVICE_MODEL).name if WHISPER_SERVICE_MODEL else "(default)"
        lang_label = "auto"
    with _print_lock:
        print(f"  [{stem}] {c('magenta', '开始识别')} [{mode_label}/{model_label}/{lang_label}/T{WHISPER_TEMPERATURE}] (文件 {file_size_mb:.1f}MB)...", flush=True)

    spinner = Spinner()
    transcribe_start = time.time()
    spinner.start(f"[{stem}] 识别中")
    try:
        if WHISPER_BACKEND == "local":
            text, retries, err = _transcribe_local(audio_file, stem, max_retries, retry_delay)
        elif WHISPER_BACKEND == "faster-whisper":
            text, retries, err = _transcribe_faster_whisper(audio_file, stem, max_retries, retry_delay)
        elif WHISPER_BACKEND == "funasr":
            text, retries, err = _transcribe_funasr(audio_file, stem, max_retries, retry_delay, timeout)
        else:
            text, retries, err = _transcribe_service(audio_file, stem, max_retries, retry_delay, timeout)
    finally:
        spinner.stop()

    if text:
        elapsed = time.time() - transcribe_start
        m, s = divmod(int(elapsed), 60)
        elapsed_str = f"{m}m{s:02d}s" if m > 0 else f"{s}s"
        with _print_lock:
            print(f"  [{stem}] {c('green', '识别完成')} ({elapsed_str}, {len(text)} 字符)", flush=True)
    else:
        log.error(f"[{stem}] 识别失败: {err}")
    return text, retries, err


def _transcribe_local(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """本地 whisper CLI 识别"""
    out_dir = audio_file.parent

    def _run():
        cmd = [
            "whisper", str(audio_file),
            "--task", WHISPER_TASK,
            "--model", WHISPER_MODEL,
            "--device", WHISPER_DEVICE,
            "--beam_size", WHISPER_BEAM_SIZE,
            "--best_of", WHISPER_BEST_OF,
            "--fp16", WHISPER_FP16,
            "--condition_on_previous_text", WHISPER_CONDITION_ON_PREV,
        ]
        if WHISPER_MODEL_DIR:
            cmd += ["--model_dir", WHISPER_MODEL_DIR]
        if WHISPER_LANGUAGE:
            cmd += ["--language", WHISPER_LANGUAGE]
        cmd += [
            "--temperature", WHISPER_TEMPERATURE,
        ]
        if WHISPER_TEMPERATURE_INC:
            cmd += ["--temperature_increment_on_fallback", WHISPER_TEMPERATURE_INC]
        if WHISPER_INITIAL_PROMPT:
            cmd += ["--initial_prompt", WHISPER_INITIAL_PROMPT]
        if WHISPER_THREADS and WHISPER_THREADS != "0":
            cmd += ["--threads", WHISPER_THREADS]
        out_ext = "json" if WHISPER_OUTPUT_FORMAT == "json" else "txt"
        cmd += [
            "--output_format", WHISPER_OUTPUT_FORMAT,
            "--output_dir", str(out_dir),
        ]

        # 合并 WHISPER_EXTRA_ARGS（CLI > .env），去重后追加
        cmd = _merge_whisper_args(cmd, _resolved_whisper_extra_args)

        # Show segment progress from stderr (whisper outputs "[00:00.000 --> 00:30.000] ...")
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True, encoding="utf-8", errors="replace",
            env=env,
        )
        import re
        last_ts = ""
        _stderr_lines = []  # 积累 stderr，失败时一并输出
        for line in proc.stderr:
            _stderr_lines.append(line.rstrip("\n"))
            line_s = line.strip()
            if not line_s:
                continue
            m = re.match(r"^\[[\d:.]+\s*-->\s*([\d:.]+)\]", line_s)
            if m:
                ts = m.group(1)
                if ts != last_ts:
                    sys.stderr.write(f"\r[{stem}] 识别中... {ts}")
                    sys.stderr.flush()
                    last_ts = ts
        if timeout > 0:
            proc.wait(timeout=timeout)
        else:
            proc.wait()
        if last_ts:
            sys.stderr.write("\n")
            sys.stderr.flush()
        _stderr_text = "\n".join(_stderr_lines).strip()
        if proc.returncode != 0:
            raise RuntimeError(
                f"whisper CLI 退出码 {proc.returncode}"
                + (f"\nstderr:\n{_stderr_text}" if _stderr_text else "")
            )
        # whisper 输出文件: {stem}.{ext}
        out_file = out_dir / f"{stem}.{out_ext}"
        if not out_file.exists():
            raise RuntimeError("whisper 输出文件未生成")
        raw = out_file.read_text(encoding="utf-8").strip()
        if WHISPER_OUTPUT_FORMAT == "json":
            import json
            try:
                return json.loads(raw).get("text", "")
            except json.JSONDecodeError:
                return raw  # 解析失败回退原文
        return raw

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        log.error(f"[{stem}] 本地 whisper 识别失败: {e}")
        return None, max_retries, str(e)[:500]


def _transcribe_faster_whisper(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """faster-whisper (CTranslate2) Python API 识别"""
    global _FW_MODEL, _FW_MODEL_CFG

    def _get_model():
        """获取或复用 WhisperModel 实例（按 model/device/compute_type 缓存）"""
        global _FW_MODEL, _FW_MODEL_CFG
        cfg = f"{WHISPER_MODEL}|{WHISPER_DEVICE}|{WHISPER_COMPUTE_TYPE}"
        if _FW_MODEL is not None and _FW_MODEL_CFG == cfg:
            return _FW_MODEL
        from faster_whisper import WhisperModel
        model_kwargs: dict = {
            "device": WHISPER_DEVICE,
            "compute_type": WHISPER_COMPUTE_TYPE,
            "num_workers": int(WHISPER_NUM_WORKERS),
        }
        if WHISPER_THREADS and WHISPER_THREADS != "0":
            model_kwargs["cpu_threads"] = int(WHISPER_THREADS)
        if WHISPER_MODEL_DIR:
            model_kwargs["download_root"] = WHISPER_MODEL_DIR
        with _print_lock:
            print(f"  [{stem}] 加载 faster-whisper 模型: {WHISPER_MODEL} (device={WHISPER_DEVICE}, compute_type={WHISPER_COMPUTE_TYPE})...", flush=True)
        _FW_MODEL = WhisperModel(WHISPER_MODEL, **model_kwargs)
        _FW_MODEL_CFG = cfg
        return _FW_MODEL

    def _run():
        model = _get_model()

        # 构建 transcribe 参数
        transcribe_kwargs: dict = {
            "beam_size": int(WHISPER_BEAM_SIZE),
            "best_of": int(WHISPER_BEST_OF),
            "temperature": float(WHISPER_TEMPERATURE),
        }
        if WHISPER_TEMPERATURE_INC:
            # faster-whisper 不支持 temperature_increment_on_fallback（原始 whisper 参数）
            # 改为构建 temperature 列表实现相同效果，如 [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]
            try:
                base_temp = float(WHISPER_TEMPERATURE)
            except (ValueError, TypeError):
                base_temp = 0.0
            try:
                inc = float(WHISPER_TEMPERATURE_INC)
            except (ValueError, TypeError):
                inc = 0.2
            if inc > 0:
                temps = [round(base_temp + i * inc, 2) for i in range(6)]
                # 去重并限制在 [0, 1] 范围内
                temps = sorted(set(max(0.0, min(1.0, t)) for t in temps))
                transcribe_kwargs["temperature"] = temps
            # inc == 0 时不做处理，使用原 base_temp 单值
        if WHISPER_LANGUAGE:
            transcribe_kwargs["language"] = WHISPER_LANGUAGE
        if WHISPER_TASK:
            transcribe_kwargs["task"] = WHISPER_TASK
        if WHISPER_INITIAL_PROMPT:
            transcribe_kwargs["initial_prompt"] = WHISPER_INITIAL_PROMPT
        transcribe_kwargs["condition_on_previous_text"] = WHISPER_CONDITION_ON_PREV.lower() in ("true", "1", "yes")

        # VAD 参数
        if WHISPER_VAD_FILTER.lower() in ("true", "1", "yes"):
            transcribe_kwargs["vad_filter"] = True
            # faster-whisper 的 VadOptions 用 threshold 而非 onset
            transcribe_kwargs["vad_parameters"] = {
                "threshold": float(WHISPER_VAD_ONSET),
            }
        else:
            transcribe_kwargs["vad_filter"] = False

        # 从 WHISPER_EXTRA_ARGS 解析额外参数（支持 CLI 覆盖）
        if _resolved_whisper_extra_args:
            transcribe_kwargs = _apply_fw_extra_args(transcribe_kwargs, _resolved_whisper_extra_args)
        # faster-whisper 不支持 temperature_increment_on_fallback，清理掉（已由上面的 WHISPER_TEMPERATURE_INC 逻辑转换为 temperature 列表）
        transcribe_kwargs.pop("temperature_increment_on_fallback", None)

        segments_iter, info = model.transcribe(str(audio_file), **transcribe_kwargs)

        # 合并所有段的文本
        segments = list(segments_iter)
        text = "".join(seg.text for seg in segments).strip()
        if not text:
            raise ValueError("faster-whisper 返回空文本")
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        import traceback as _tb
        _audio_size = audio_file.stat().st_size if audio_file.exists() else "N/A"
        _err_details = (
            f"[{stem}] faster-whisper 识别失败\n"
            f"  audio : {audio_file} ({_audio_size} bytes)\n"
            f"  model : {WHISPER_MODEL} (device={WHISPER_DEVICE}, compute_type={WHISPER_COMPUTE_TYPE})\n"
            f"  model_dir: {WHISPER_MODEL_DIR or '~/.cache/huggingface/hub'}\n"
            f"  error : {e}\n"
        )
        # 如果是 SSL/cert 相关错误，给出解决提示
        _err_str = str(e)
        if any(k in _err_str.lower() for k in ("certificate", "ssl", "CERTIFICATE_VERIFY_FAILED", "ConnectError")):
            _err_details += (
                f"  ⚠️ 疑似 SSL/证书错误，可能的解决方案：\n"
                f"    1. 设置环境变量 SSL_CERT_FILE=/path/to/cert.pem\n"
                f"    2. pip install --upgrade certifi\n"
                f"    3. 设置 WHISPER_MODEL_DIR 指向已下载的本地模型目录（跳过在线下载）\n"
            )
        log.error(_err_details + f"  traceback:\n{_tb.format_exc()}")
        return None, max_retries, _err_details.strip()


def _apply_fw_extra_args(base_kwargs: dict, extra_args: list[str]) -> dict:
    """将 WHISPER_EXTRA_ARGS 中的额外参数应用到 faster-whisper transcribe kwargs。

    支持两种格式：
      --key value         → kwargs[key] = value (自动转 bool/int/float)
      --key               → kwargs[key] = True (布尔标志)
    对已有参数进行覆盖。
    """
    if not extra_args:
        return base_kwargs
    kwargs = dict(base_kwargs)
    i = 0
    while i < len(extra_args):
        arg = extra_args[i]
        if arg.startswith("--"):
            key = arg[2:]
            if i + 1 < len(extra_args) and not extra_args[i + 1].startswith("--"):
                val = extra_args[i + 1]
                # 自动类型转换
                if val.lower() in ("true", "1", "yes"):
                    kwargs[key] = True
                elif val.lower() in ("false", "0", "no"):
                    kwargs[key] = False
                elif "." in val:
                    try:
                        kwargs[key] = float(val)
                    except ValueError:
                        kwargs[key] = val
                else:
                    try:
                        kwargs[key] = int(val)
                    except ValueError:
                        kwargs[key] = val
                i += 2
            else:
                # 无值 flag → True
                kwargs[key] = True
                i += 1
        else:
            i += 1
    return kwargs


def _transcribe_funasr(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR 识别（cli 模式走 AutoModel, service 模式走 HTTP 转发到 funasr-server）"""
    if FUNASR_MODE == "service":
        return _transcribe_funasr_service(audio_file, stem, max_retries, retry_delay, timeout)
    return _transcribe_funasr_cli(audio_file, stem, max_retries, retry_delay, timeout)


def _transcribe_funasr_cli(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR CLI 模式：本地 AutoModel 推理（首次自动下载 ModelScope 模型）"""
    global _FUNASR_MODEL, _FUNASR_MODEL_CFG

    def _get_model():
        """获取或复用 AutoModel 实例（按 model/device/vad/punc/spk 组合缓存）"""
        global _FUNASR_MODEL, _FUNASR_MODEL_CFG
        cfg = f"{FUNASR_MODEL}|{FUNASR_DEVICE}|{FUNASR_VAD_MODEL}|{FUNASR_PUNC_MODEL}|{FUNASR_SPK_MODEL}|{FUNASR_EMOTION_MODEL}"
        if _FUNASR_MODEL is not None and _FUNASR_MODEL_CFG == cfg:
            return _FUNASR_MODEL
        from funasr import AutoModel  # 延迟导入，避免 funasr 未装时启动失败
        model_kwargs: dict = {
            "model": FUNASR_MODEL,
            "device": FUNASR_DEVICE,
        }
        # 可选辅助模型（空字符串=不启用）
        if FUNASR_VAD_MODEL:
            model_kwargs["vad_model"] = FUNASR_VAD_MODEL
        if FUNASR_PUNC_MODEL:
            model_kwargs["punc_model"] = FUNASR_PUNC_MODEL
        if FUNASR_SPK_MODEL:
            model_kwargs["spk_model"] = FUNASR_SPK_MODEL
        if FUNASR_EMOTION_MODEL:
            model_kwargs["emotion_model"] = FUNASR_EMOTION_MODEL
        with _print_lock:
            _models = [FUNASR_MODEL]
            if FUNASR_VAD_MODEL:    _models.append(FUNASR_VAD_MODEL)
            if FUNASR_PUNC_MODEL:   _models.append(FUNASR_PUNC_MODEL)
            if FUNASR_SPK_MODEL:    _models.append(FUNASR_SPK_MODEL)
            if FUNASR_EMOTION_MODEL: _models.append(FUNASR_EMOTION_MODEL)
            print(f"  [{stem}] 加载 FunASR 模型: {'+'.join(_models)} (device={FUNASR_DEVICE})...", flush=True)
        _FUNASR_MODEL = AutoModel(**model_kwargs)
        _FUNASR_MODEL_CFG = cfg
        return _FUNASR_MODEL

    def _run():
        model = _get_model()

        # 构建 generate 参数
        generate_kwargs: dict = {
            "input": str(audio_file),
            "batch_size_s": int(FUNASR_BATCH_SIZE_S),
        }
        if FUNASR_HOTWORD:
            generate_kwargs["hotword"] = FUNASR_HOTWORD
        if FUNASR_LANGUAGE:
            generate_kwargs["language"] = FUNASR_LANGUAGE
        if FUNASR_VAD_MODEL and FUNASR_VAD_MAX_SEGMENT and FUNASR_VAD_MAX_SEGMENT != "0":
            generate_kwargs["vad_kwargs"] = {"max_single_segment_time": int(FUNASR_VAD_MAX_SEGMENT)}

        # 从 FUNASR_EXTRA_ARGS 解析额外参数（支持 CLI 覆盖）
        if _resolved_funasr_extra_args:
            generate_kwargs = _apply_fw_extra_args(generate_kwargs, _resolved_funasr_extra_args)

        results = model.generate(**generate_kwargs)
        if not results:
            raise ValueError("FunASR 返回空结果")
        # results[0] 通常是 dict，含 "text" 字段
        first = results[0]
        if isinstance(first, dict):
            text = (first.get("text") or "").strip()
        else:
            text = str(first).strip()
        if not text:
            raise ValueError("FunASR 返回空文本")
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        import traceback as _tb
        _audio_size = audio_file.stat().st_size if audio_file.exists() else "N/A"
        _err_details = (
            f"[{stem}] FunASR CLI 识别失败\n"
            f"  audio    : {audio_file} ({_audio_size} bytes)\n"
            f"  mode     : cli (AutoModel)\n"
            f"  model    : {FUNASR_MODEL} (device={FUNASR_DEVICE})\n"
            f"  vad/punc : {FUNASR_VAD_MODEL} / {FUNASR_PUNC_MODEL}\n"
            f"  hotword  : {FUNASR_HOTWORD or '(none)'}\n"
            f"  error    : {e}\n"
        )
        # 如果是 SSL/cert 相关错误，给出解决提示
        _err_str = str(e)
        if any(k in _err_str.lower() for k in ("certificate", "ssl", "CERTIFICATE_VERIFY_FAILED", "ConnectError")):
            _err_details += (
                f"  ⚠️ 疑似 SSL/证书错误，可能的解决方案：\n"
                f"    1. 设置环境变量 SSL_CERT_FILE=/path/to/cert.pem\n"
                f"    2. pip install --upgrade certifi\n"
                f"    3. 设置 FUNASR_MODEL_DIR 指向已下载的本地模型目录（跳过在线下载）\n"
            )
        log.error(_err_details + f"  traceback:\n{_tb.format_exc()}")
        return None, max_retries, _err_details.strip()


def _transcribe_funasr_service(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR 服务模式：调用 funasr-server OpenAI 兼容 API。"""
    global _Service_model_loaded

    def _run():
        # ── 按需切换模型（仅当指定了 model 时，且服务支持动态加载） ──
        if FUNASR_SERVICE_MODEL and FUNASR_SERVICE_MODEL != _Service_model_loaded:
            with _print_lock:
                print(f"  [{stem}] 加载 FunASR 服务模型: {FUNASR_SERVICE_MODEL}", flush=True)

        with open(audio_file, "rb") as f:
            data = {
                "model": FUNASR_SERVICE_MODEL,
                "response_format": "json",
            }
            if FUNASR_HOTWORD:
                # OpenAI 兼容 API 的 hotword 一般以 prompt 形式传递
                data["prompt"] = FUNASR_HOTWORD
            resp = requests.post(
                f"{FUNASR_SERVICE_URL}/v1/audio/transcriptions",
                files={"file": (audio_file.name, f, "audio/wav")},
                data=data,
                timeout=timeout if timeout > 0 else None,
            )
        resp.raise_for_status()
        data = resp.json()
        text = (data.get("text") or "").strip()
        if not text:
            raise ValueError("FunASR 服务返回空文本")
        _Service_model_loaded = FUNASR_SERVICE_MODEL
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        log.error(f"[{stem}] FunASR 服务识别失败: {e}")
        return None, max_retries, str(e)[:500]


def _transcribe_funasr(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR 识别（cli 模式走 AutoModel, service 模式走 HTTP 转发到 funasr-server）"""
    if FUNASR_MODE == "service":
        return _transcribe_funasr_service(audio_file, stem, max_retries, retry_delay, timeout)
    return _transcribe_funasr_cli(audio_file, stem, max_retries, retry_delay, timeout)


def _transcribe_funasr_cli(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR CLI 模式：本地 AutoModel 推理（首次自动下载 ModelScope 模型）"""
    global _FUNASR_MODEL, _FUNASR_MODEL_CFG

    def _get_model():
        """获取或复用 AutoModel 实例（按 model/device/vad/punc/spk 组合缓存）"""
        global _FUNASR_MODEL, _FUNASR_MODEL_CFG
        cfg = f"{FUNASR_MODEL}|{FUNASR_DEVICE}|{FUNASR_VAD_MODEL}|{FUNASR_PUNC_MODEL}|{FUNASR_SPK_MODEL}|{FUNASR_EMOTION_MODEL}"
        if _FUNASR_MODEL is not None and _FUNASR_MODEL_CFG == cfg:
            return _FUNASR_MODEL
        from funasr import AutoModel  # 延迟导入，避免 funasr 未装时启动失败
        model_kwargs: dict = {
            "model": FUNASR_MODEL,
            "device": FUNASR_DEVICE,
        }
        # 可选辅助模型（空字符串=不启用）
        if FUNASR_VAD_MODEL:
            model_kwargs["vad_model"] = FUNASR_VAD_MODEL
        if FUNASR_PUNC_MODEL:
            model_kwargs["punc_model"] = FUNASR_PUNC_MODEL
        if FUNASR_SPK_MODEL:
            model_kwargs["spk_model"] = FUNASR_SPK_MODEL
        if FUNASR_EMOTION_MODEL:
            model_kwargs["emotion_model"] = FUNASR_EMOTION_MODEL
        with _print_lock:
            _models = [FUNASR_MODEL]
            if FUNASR_VAD_MODEL:    _models.append(FUNASR_VAD_MODEL)
            if FUNASR_PUNC_MODEL:   _models.append(FUNASR_PUNC_MODEL)
            if FUNASR_SPK_MODEL:    _models.append(FUNASR_SPK_MODEL)
            if FUNASR_EMOTION_MODEL: _models.append(FUNASR_EMOTION_MODEL)
            print(f"  [{stem}] 加载 FunASR 模型: {'+'.join(_models)} (device={FUNASR_DEVICE})...", flush=True)
        _FUNASR_MODEL = AutoModel(**model_kwargs)
        _FUNASR_MODEL_CFG = cfg
        return _FUNASR_MODEL

    def _run():
        model = _get_model()

        # 构建 generate 参数
        generate_kwargs: dict = {
            "input": str(audio_file),
            "batch_size_s": int(FUNASR_BATCH_SIZE_S),
        }
        if FUNASR_HOTWORD:
            generate_kwargs["hotword"] = FUNASR_HOTWORD
        if FUNASR_LANGUAGE:
            generate_kwargs["language"] = FUNASR_LANGUAGE
        if FUNASR_VAD_MODEL and FUNASR_VAD_MAX_SEGMENT and FUNASR_VAD_MAX_SEGMENT != "0":
            generate_kwargs["vad_kwargs"] = {"max_single_segment_time": int(FUNASR_VAD_MAX_SEGMENT)}

        # 从 FUNASR_EXTRA_ARGS 解析额外参数（支持 CLI 覆盖）
        if _resolved_funasr_extra_args:
            generate_kwargs = _apply_fw_extra_args(generate_kwargs, _resolved_funasr_extra_args)

        results = model.generate(**generate_kwargs)
        if not results:
            raise ValueError("FunASR 返回空结果")
        # results[0] 通常是 dict，含 "text" 字段
        first = results[0]
        if isinstance(first, dict):
            text = (first.get("text") or "").strip()
        else:
            text = str(first).strip()
        if not text:
            raise ValueError("FunASR 返回空文本")
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        import traceback as _tb
        _audio_size = audio_file.stat().st_size if audio_file.exists() else "N/A"
        _err_details = (
            f"[{stem}] FunASR CLI 识别失败\n"
            f"  audio    : {audio_file} ({_audio_size} bytes)\n"
            f"  mode     : cli (AutoModel)\n"
            f"  model    : {FUNASR_MODEL} (device={FUNASR_DEVICE})\n"
            f"  vad/punc : {FUNASR_VAD_MODEL} / {FUNASR_PUNC_MODEL}\n"
            f"  hotword  : {FUNASR_HOTWORD or '(none)'}\n"
            f"  error    : {e}\n"
        )
        # 如果是 SSL/cert 相关错误，给出解决提示
        _err_str = str(e)
        if any(k in _err_str.lower() for k in ("certificate", "ssl", "CERTIFICATE_VERIFY_FAILED", "ConnectError")):
            _err_details += (
                f"  ⚠️ 疑似 SSL/证书错误，可能的解决方案：\n"
                f"    1. 设置环境变量 SSL_CERT_FILE=/path/to/cert.pem\n"
                f"    2. pip install --upgrade certifi\n"
                f"    3. 设置 FUNASR_MODEL_DIR 指向已下载的本地模型目录（跳过在线下载）\n"
            )
        log.error(_err_details + f"  traceback:\n{_tb.format_exc()}")
        return None, max_retries, _err_details.strip()


def _transcribe_funasr_service(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """FunASR 服务模式：调用 funasr-server OpenAI 兼容 API。"""
    global _Service_model_loaded

    def _run():
        # ── 按需切换模型（仅当指定了 model 时，且服务支持动态加载） ──
        if FUNASR_SERVICE_MODEL and FUNASR_SERVICE_MODEL != _Service_model_loaded:
            with _print_lock:
                print(f"  [{stem}] 加载 FunASR 服务模型: {FUNASR_SERVICE_MODEL}", flush=True)

        with open(audio_file, "rb") as f:
            data = {
                "model": FUNASR_SERVICE_MODEL,
                "response_format": "json",
            }
            if FUNASR_HOTWORD:
                # OpenAI 兼容 API 的 hotword 一般以 prompt 形式传递
                data["prompt"] = FUNASR_HOTWORD
            resp = requests.post(
                f"{FUNASR_SERVICE_URL}/v1/audio/transcriptions",
                files={"file": (audio_file.name, f, "audio/wav")},
                data=data,
                timeout=timeout if timeout > 0 else None,
            )
        resp.raise_for_status()
        data = resp.json()
        text = (data.get("text") or "").strip()
        if not text:
            raise ValueError("FunASR 服务返回空文本")
        _Service_model_loaded = FUNASR_SERVICE_MODEL
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        log.error(f"[{stem}] FunASR 服务识别失败: {e}")
        return None, max_retries, str(e)[:500]


def _transcribe_service(
    audio_file: Path, stem: str,
    max_retries: int, retry_delay: float,
    timeout: int = 0,
) -> tuple[str | None, int, str | None]:
    """远程 whisper.cpp server 识别（无内建进度显示，由调用方 Spinner 负责）"""
    global _SERVICE_MODEL_LOADED

    def _run():
        # ── 按需切换模型（/load） ──
        if WHISPER_SERVICE_MODEL and WHISPER_SERVICE_MODEL != _SERVICE_MODEL_LOADED:
            with _print_lock:
                print(f"  [{stem}] 切换模型: {WHISPER_SERVICE_MODEL}", flush=True)
            resp = requests.post(
                f"{WHISPER_SERVICE}/load",
                data={"model": WHISPER_SERVICE_MODEL},
                timeout=30,
            )
            resp.raise_for_status()
            _SERVICE_MODEL_LOADED = WHISPER_SERVICE_MODEL

        # ── 语音识别（/inference） ──
        with open(audio_file, "rb") as f:
            data = {
                "temperature": WHISPER_TEMPERATURE,
                "temperature_inc": WHISPER_TEMPERATURE_INC,
                "response_format": WHISPER_OUTPUT_FORMAT,
            }
            resp = requests.post(
                f"{WHISPER_SERVICE}/inference",
                files={"file": (audio_file.name, f, "audio/wav")},
                data=data,
                timeout=timeout if timeout > 0 else None,
            )
        resp.raise_for_status()
        data = resp.json()
        text = data.get("text", "").strip()
        if not text:
            raise ValueError("whisper 返回空文本")
        return text

    try:
        text, retries_used, err = retry_call(
            _run, max_retries=max_retries, base_delay=retry_delay, task_label=stem,
        )
        if err:
            return None, retries_used, err
        return text, 0, None
    except Exception as e:
        log.error(f"[{stem}] 服务 whisper 识别失败: {e}")
        return None, max_retries, str(e)[:500]


# ─────────────────────────────── 增量进度写回 ────────────────────────────────

def save_task_progress(result: TaskResult) -> None:
    """每完成一个任务立即写入 progress JSON，方便用户随时查看中间结果。

    目录结构: output/progress/{sheet}/task_{stem}.json
    文件包含 content（识别文本）和 keywords（AI 分析）等关键字段。
    """
    progress_dir = PROGRESS_DIR / result.sheet
    progress_dir.mkdir(parents=True, exist_ok=True)
    progress_file = progress_dir / f"task_{result.stem}.json"

    # 提取 content 和 keywords（StepResult.file 被借用存文本）
    content = result.transcribe.file if (
        result.transcribe.status == "success" and isinstance(result.transcribe.file, str)
    ) else None
    keywords = result.analyze.file if (
        result.analyze.status == "success" and isinstance(result.analyze.file, str)
    ) else None

    data = {
        "sheet": result.sheet,
        "id_val": result.id_val,
        "title": result.title,
        "stem": result.stem,
        "platform": result.platform,
        "video_url": result.video_url,
        "overall_status": result.overall_status,
        "error": result.error,
        "content": content,
        "keywords": keywords,
        "download": {
            "status": result.download.status,
            "file": result.download.file,
            "error": result.download.error,
            "retries_used": result.download.retries_used,
        },
        "transcode": {
            "status": result.transcode.status,
            "file": result.transcode.file,
            "error": result.transcode.error,
            "retries_used": result.transcode.retries_used,
        },
        "transcribe": {
            "status": result.transcribe.status,
            "error": result.transcribe.error,
            "retries_used": result.transcribe.retries_used,
        },
        "analyze": {
            "status": result.analyze.status,
            "error": result.analyze.error,
            "retries_used": result.analyze.retries_used,
        },
        "timestamp": datetime.now().isoformat(),
    }

    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 打印简短提示（使用动态 PROGRESS_DIR 而非硬编码路径）
    rel_path = os.path.relpath(str(progress_file), str(BASE_DIR)).replace("\\", "/")
    info_parts = [result.overall_status]
    if content:
        info_parts.append(f"content({len(content)}字符)")
    if keywords:
        info_parts.append(f"keywords({len(keywords)}字符)")
    with _print_lock:
        print(
            c("dim", f"  📄 进度已保存: {rel_path}")
            + f"  [{', '.join(info_parts)}]",
            flush=True,
        )

    # 释放内存：content / keywords 已持久化到 JSON，可安全清空
    result.transcribe.file = None
    result.analyze.file = None

    # ── 断点续跑：实时回写 Excel（断电时仍能保留已完成的识别/分析结果）──
    if content:
        try:
            write_excel_cell(result.sheet, str(result.id_val), COL_CONTENT, content)
        except Exception as e:
            log.warning(f"[{result.stem}] 实时写 Excel content 失败（不影响 progress JSON）: {e}")
    if keywords:
        try:
            write_excel_cell(result.sheet, str(result.id_val), COL_KEYWORDS, keywords)
        except Exception as e:
            log.warning(f"[{result.stem}] 实时写 Excel keywords 失败（不影响 progress JSON）: {e}")


# ─────────────────────────────── Excel 实时写回 ─────────────────────────────

def write_excel_cell(sheet_name: str, key: str, col_name: str, value: str) -> bool:
    """单条 Excel 单元格写回（断点续跑场景：每条任务完成立即写）。

    使用 _excel_lock 串行化，openpyxl 非线程安全。
    返回是否成功写入。
    """
    if not value or not value.strip():
        return False
    with _excel_lock:
        try:
            wb = load_workbook(str(EXCEL_FILE))
        except Exception as e:
            log.error(f"打开 Excel 失败: {e}")
            return False
        try:
            if sheet_name not in wb.sheetnames:
                log.warning(f"Sheet [{sheet_name}] 不存在，跳过写入")
                return False
            ws = wb[sheet_name]
            headers = {cell.value: cell.column for cell in ws[1]}
            if col_name not in headers:
                log.warning(f"[{sheet_name}] 找不到 {col_name} 列，跳过写入")
                return False
            target_col = headers[col_name]
            id_col = headers.get(COL_ID)
            title_col = headers.get(COL_TITLE)

            matched = False
            for row in ws.iter_rows(min_row=2):
                id_val = row[id_col - 1].value if id_col else None
                title_val = row[title_col - 1].value if title_col else None
                if id_col and id_val is not None:
                    try:
                        if str(int(float(id_val))) == str(key):
                            matched = True
                    except (ValueError, TypeError):
                        pass
                if not matched and title_col:
                    if str(title_val) == str(key):
                        matched = True
                if matched:
                    # Excel 单元格字符上限 32767，超出需截断
                    EXCEL_MAX_CHARS = 32767
                    safe_value = value[:EXCEL_MAX_CHARS] if len(value) > EXCEL_MAX_CHARS else value
                    if len(value) > EXCEL_MAX_CHARS:
                        log.warning(f"[{sheet_name}/{key}] {col_name} 截断 {len(value)} -> {EXCEL_MAX_CHARS} 字符 (Excel 限制)")
                    row[target_col - 1].value = safe_value
                    return True

            log.warning(f"[{sheet_name}] 未找到匹配行 key={key}")
            return False
        finally:
            try:
                wb.save(str(EXCEL_FILE))
            except Exception as e:
                log.error(f"保存 Excel 失败: {e}")
            wb.close()


def write_all_contents_to_excel(results: list[TaskResult], keywords_dict: dict[tuple[str, str], str] | None = None, content_dict: dict[tuple[str, str], str] | None = None):
    """
    将所有识别文本批量写回 Excel。
    使用 openpyxl 直接操作，单线程安全。
    通过 content_dict 直接传入内容（避免依赖 result 对象上可能已被释放的大文本字段）。
    """
    if not results:
        return

    # 收集需要写入的数据
    if content_dict is not None:
        updates = content_dict
    else:
        updates: dict[tuple[str, str], str] = {}
        for tr in results:
            if tr.transcribe.status == "success" and tr.transcribe.file:
                text = tr.transcribe.file
                if text.strip():
                    updates[(tr.sheet, tr.id_val)] = text

    if not updates:
        return

    log.info(f"批量写入 {len(updates)} 条识别文本到 Excel...")
    for (sheet_name, key), text in updates.items():
        if write_excel_cell(sheet_name, str(key), COL_CONTENT, text):
            log.info(f"[{sheet_name}/{key}] content 已写入（{len(text)} 字符）")

    # 写入 keywords 列（AI 分析结果）
    if keywords_dict:
        for (kw_sheet, kw_key), kw_text in keywords_dict.items():
            if write_excel_cell(kw_sheet, str(kw_key), COL_KEYWORDS, kw_text):
                log.info(f"[{kw_sheet}/{kw_key}] keywords 已写入（{len(kw_text)} 字符）")
    log.info("Excel 写入完成")


# ─────────────────────────────── 报告生成 ───────────────────────────────────

def compute_summary(results: list[TaskResult]) -> dict:
    """统一计算 result 的 status 统计，避免重复遍历。"""
    success = partial = failed = no_video = 0
    for r in results:
        s = r.overall_status
        if s == "success":
            success += 1
        elif s == "partial":
            partial += 1
        elif s == "failed":
            failed += 1
        elif s == "no_video":
            no_video += 1
    return {"total": len(results), "success": success,
            "partial": partial, "failed": failed, "no_video": no_video}


def _check_and_confirm_env(steps: list[str], dry_run: bool, confirm_msg: str) -> bool:
    """环境检测 + 用户确认（统一逻辑，消除 run() 和 run_from_report() 的重复代码）。

    - 非 dry_run：检测环境，有问题则列出并询问用户是否继续；用户取消返回 False
    - dry_run：检测环境并打印详细状态，始终返回 True（干跑不执行）
    返回 True = 继续执行，False = 用户取消
    """
    env = check_environment(steps)

    if dry_run:
        # 干跑模式：打印环境检测详情
        print("\n  --- 环境检测 ---")
        # yt-dlp
        if "download" in steps:
            if env["ytdlp"]:
                print(f"  ✅ yt-dlp: 可用 ({YTDLP})")
            else:
                print(f"  ❌ yt-dlp: 不可用 ({YTDLP})")
        else:
            print(f"  ⏭ yt-dlp: 未启用（步骤不含 download）")
        # ffmpeg
        if "transcode" in steps:
            if env["ffmpeg"]:
                print(f"  ✅ ffmpeg: 可用 ({FFMPEG})")
            else:
                print(f"  ❌ ffmpeg: 不可用 ({FFMPEG})")
        else:
            print(f"  ⏭ ffmpeg: 未启用（步骤不含 transcode）")
        # ffprobe
        if "transcode" in steps:
            if env["ffprobe"]:
                print(f"  ✅ ffprobe: 可用 ({FFPROBE})")
            else:
                print(f"  ❌ ffprobe: 不可用 ({FFPROBE})")
        else:
            print(f"  ⏭ ffprobe: 未启用（步骤不含 transcode）")
        # whisper
        if "transcribe" in steps:
            if WHISPER_BACKEND == "local":
                backend_info = "本地 whisper CLI"
            elif WHISPER_BACKEND == "faster-whisper":
                backend_info = "faster-whisper (faster_whisper)"
            elif WHISPER_BACKEND == "funasr":
                if FUNASR_MODE == "service":
                    backend_info = f"funasr/service ({FUNASR_SERVICE_URL})"
                else:
                    backend_info = f"funasr/cli ({FUNASR_MODEL})"
            else:
                backend_info = f"service {WHISPER_SERVICE}"
            if env["whisper"]:
                print(f"  ✅ whisper ({backend_info}): 可用")
            else:
                print(f"  ❌ whisper ({backend_info}): 不可用")
        else:
            print(f"  ⏭ whisper: 未启用（步骤不含 transcribe）")
        # AI 分析
        if "analyze" in steps:
            ai_model = os.getenv("AI_MODEL", "")
            if env["ai"]:
                print(f"  ✅ AI分析 ({ai_model}): 配置完整")
            else:
                print(f"  ❌ AI分析: {env['issues'][-1] if env['issues'] else '配置不完整'}")
        else:
            print(f"  ⏭ AI分析: 未启用（步骤不含 analyze）")
        return True

    # 非 dry_run：预检并询问
    if not env["all_ok"]:
        print("\n" + "=" * 60)
        print("  ⚠️  工具/服务预检：以下依赖不可用")
        print("=" * 60)
        for issue in env["issues"]:
            print(f"  • {issue}")
        print("\n  涉及的步骤将失败。")
        try:
            choice = input(f"\n  {confirm_msg} (输入 'yes' 继续，其他任意键取消): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            choice = "no"
        if choice != "yes":
            log.info("用户取消执行（工具不可用）")
            return False

    return True


def generate_report(results: list[TaskResult], config: dict, sheet_name: str | None = None) -> Path | list[Path]:
    """生成执行报告 JSON 文件。
    - sheet_name 提供时：报告存入 REPORTS_DIR/{sheet_name}/report_{ts}.json
    - 不提供时：按 r.sheet 分组，每 sheet 调用自身，返回路径列表
    """
    if sheet_name is None:
        # ── 按 sheet 分组生成 ──
        sheet_groups: dict[str, list[TaskResult]] = {}
        for r in results:
            sheet_groups.setdefault(r.sheet, []).append(r)
        paths = []
        for sheet, items in sheet_groups.items():
            paths.append(generate_report(items, config, sheet))
        return paths

    # ── 单 sheet 报告 ──
    dir_path = REPORTS_DIR / sheet_name
    dir_path.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = dir_path / f"report_{timestamp}.json"

    summary = compute_summary(results)

    report = {
        "timestamp": datetime.now().isoformat(),
        "config": config,
        "summary": summary,
        "items": [r.to_dict() for r in results],
        "failed_items": [
            {
                "sheet": r.sheet,
                "id": r.id_val,
                "title": r.title,
                "stem": r.stem,
                "error": r.error,
                "download_error": r.download.error,
                "transcode_error": r.transcode.error,
                "transcribe_error": r.transcribe.error,
            }
            for r in results if r.overall_status in ("failed", "partial")
        ],
    }

    with open(report_file, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    log.info(f"报告已生成: {report_file}")
    return report_file


def print_report_summary(results: list[TaskResult]):
    """打印控制台摘要"""
    s = compute_summary(results)
    success = s["success"]
    partial = s["partial"]
    failed = s["failed"]
    no_vid = s["no_video"]

    print(f"\n{c('dim', '=' * 60)}")
    print(c("bold", f"  执行摘要"))
    print(c("dim", "=" * 60))
    print(f"  总计: {len(results)}")
    print(f"  {c('green', '✅ 成功')}: {success}")
    print(f"  {c('yellow', '⚠️ 部分成功')}: {partial}")
    print(f"  {c('red', '❌ 失败')}: {failed}")
    print(f"  {c('dim', '⏭️ 无视频ID')}: {no_vid}")
    print(c("dim", "=" * 60))

    # 列出所有非成功项
    failures = [r for r in results if r.overall_status != "success"]
    if failures:
        print(c("red", f"\n失败/异常详情:"))
        for r in failures:
            icon = {"partial": "⚠️", "failed": "❌", "no_video": "⏭️"}.get(r.overall_status, "?")
            print(f"  {icon} [{r.sheet}] {r.id_val} ({r.title[:30] if r.title else 'N/A'})")
            if r.error:
                _print_long(r.error, max_chars=800, indent="       ")
            if r.download.status == "failed":
                _print_long(r.download.error, max_chars=800, indent="       ")
            if r.transcode.status == "failed":
                _print_long(r.transcode.error, max_chars=800, indent="       ")
            if r.transcribe.status == "failed":
                _print_long(r.transcribe.error, max_chars=800, indent="       ")


# ─────────────────────────────── 单任务处理 ─────────────────────────────────

def process_one_task(
    row: pd.Series, sheet_name: str, steps: list[str],
    max_retries: int, retry_delay: float, force: bool,
    whisper_available: bool,
    position_label: str = "",
    download_timeout: int = 1800,
    transcode_timeout: int = 1200,
    transcribe_timeout: int = 0,
    analyze_timeout: int = 300,
) -> TaskResult:
    """处理单个视频的全流程（在独立线程中执行）"""
    pkey, vid = get_video_id(row)
    stem = stem_name(row, sheet_name)
    key = row_key(row)
    title = str(row.get(COL_TITLE, ""))
    url = build_url(pkey, vid) if pkey else None

    result = TaskResult(
        sheet=sheet_name, id_val=key, title=title,
        platform=pkey, video_url=url, stem=stem,
    )

    # 打印任务头部（带分隔线）
    with _print_lock:
        print()
        print(c("dim", "─" * 62))
        print(
            c("bold", f"  ▶ Task {position_label or '?'}")
            + "  " + c("dim", f"[{stem}]  sheet={sheet_name}  platform={pkey or 'N/A'}")
        )
        if title:
            print(c("dim", f"  title: {title[:50]}"))
        print(c("dim", "─" * 62))
    log.info(f"[{stem}] 开始处理 (sheet={sheet_name}, platform={pkey or 'N/A'})")

    # ── 断点续跑：读取上次 progress，按 status=success 跳过已完成 step ──
    # 规则：success 必须有产物可校验（文件存在 / 内容长度达标），否则视为未完成
    prior = None if force else load_task_progress(sheet_name, stem)
    skip_steps: set[str] = set()
    if prior:
        # download：上次成功时记录了文件路径，校验文件仍存在
        if prior.get("download", {}).get("status") == "success":
            f = prior["download"].get("file")
            if f and Path(f).exists() and Path(f).stat().st_size > 0:
                skip_steps.add("download")
                result.download = StepResult("success", file=f)
        # transcode：校验转码文件
        if prior.get("transcode", {}).get("status") == "success":
            f = prior["transcode"].get("file")
            if f and Path(f).exists() and Path(f).stat().st_size > 0:
                skip_steps.add("transcode")
                result.transcode = StepResult("success", file=f)
        # transcribe：校验 transcript 文本文件
        if prior.get("transcribe", {}).get("status") == "success":
            tp = transcript_path(sheet_name, stem)
            if tp.exists():
                try:
                    cached_text = tp.read_text(encoding="utf-8")
                except OSError:
                    cached_text = ""
                ok, _ = validate_transcript_text(cached_text)
                if ok:
                    skip_steps.add("transcribe")
                    result.transcribe = StepResult("success", file=cached_text)
        # analyze：校验 keywords 文本文件
        if prior.get("analyze", {}).get("status") == "success":
            kp = keywords_path(sheet_name, stem)
            if kp.exists():
                try:
                    cached_kw = kp.read_text(encoding="utf-8")
                except OSError:
                    cached_kw = ""
                ok, _ = validate_keywords_text(cached_kw)
                if ok:
                    skip_steps.add("analyze")
                    result.analyze = StepResult("success", file=cached_kw)
        if skip_steps:
            with _print_lock:
                print(
                    c("cyan", f"  ♻️ 断点续跑：跳过已完成步骤 {sorted(skip_steps)}")
                    + c("dim", f"  [来源: progress JSON]"),
                    flush=True,
                )

    # ── 下载 ──
    if "download" in steps:
        if not pkey:
            result.download = StepResult("skipped")
            result.overall_status = "no_video"
            result.error = "无可用视频 ID"
            log.warning(f"[{stem}] 无可用视频 ID，标记为 no_video")
            return result

        if "download" in skip_steps:
            dl_file = Path(result.download.file)
            with _print_lock:
                print(c("dim", f"  [{stem}] ♻️ 跳过 download，复用 {dl_file.name}"), flush=True)
        else:
            try:
                dl_file, retries, err = step_download(row, sheet_name, max_retries, retry_delay, force, download_timeout)
            except Exception as e:
                dl_file, retries, err = None, max_retries, str(e)[:500]

            # 失败时清理残留（不留半成功文件）
            if not dl_file:
                dl_dir = DOWNLOADS_DIR / sheet_name
                _cleanup_partial_files(dl_dir, stem)

            result.download = StepResult(
                status="success" if dl_file else "failed",
                file=str(dl_file) if dl_file else None,
                error=err,
                retries_used=retries,
            )

            if not dl_file:
                result.overall_status = "failed"
                result.error = f"下载失败: {err}"
                return result
    else:
        dl_dir = DOWNLOADS_DIR / sheet_name
        dl_file = find_downloaded_file(dl_dir, stem)
        if dl_file:
            result.download = StepResult("success", file=str(dl_file))
        else:
            result.download = StepResult("skipped")

    # ── 转码 ──
    if "transcode" in steps and dl_file:
        if "transcode" in skip_steps:
            tc_file = Path(result.transcode.file)
            with _print_lock:
                print(c("dim", f"  [{stem}] ♻️ 跳过 transcode，复用 {tc_file.name}"), flush=True)
        else:
            try:
                tc_file, retries, err = step_transcode(dl_file, sheet_name, max_retries, retry_delay, force, transcode_timeout)
            except Exception as e:
                tc_file, retries, err = None, max_retries, str(e)[:500]

            # 失败时清理损坏的转码文件
            if not tc_file:
                tc_dir = TRANSCODED_DIR / sheet_name
                bad = tc_dir / (stem + TRANSCODE_EXT)
                if bad.exists() and (bad.stat().st_size == 0 or not force):
                    safe_remove(bad)

            result.transcode = StepResult(
                status="success" if tc_file else "failed",
                file=str(tc_file) if tc_file else None,
                error=err,
                retries_used=retries,
            )

            if not tc_file:
                result.overall_status = "partial"
                result.error = f"下载成功但转码失败: {err}"
                return result
    else:
        tc_dir = TRANSCODED_DIR / sheet_name
        tc_path = tc_dir / (stem + TRANSCODE_EXT)
        if tc_path.exists():
            result.transcode = StepResult("success", file=str(tc_path))
            tc_file = tc_path
        else:
            result.transcode = StepResult("skipped")
            tc_file = None

    # ── 识别 ──
    if "transcribe" in steps and tc_file:
        if not whisper_available:
            result.transcribe = StepResult("failed", error=f"whisper 服务不可达 ({WHISPER_SERVICE})")
            result.overall_status = "partial"
            result.error = "下载+转码成功但 whisper 服务不可达"
            return result

        if "transcribe" in skip_steps:
            text = result.transcribe.file
            with _print_lock:
                print(c("dim", f"  [{stem}] ♻️ 跳过 transcribe，复用缓存文本({len(text)}字符)"), flush=True)
        else:
            try:
                text, retries, err = step_transcribe(tc_file, max_retries, retry_delay, transcribe_timeout)
            except Exception as e:
                text, retries, err = None, max_retries, str(e)[:500]

            # 校验：transcribe 不允许"半成功"——产物不合格即清理
            ok, validate_err = validate_transcript_text(text)
            if not ok:
                text = None
                err = err or validate_err
                safe_remove(transcript_path(sheet_name, stem))

            result.transcribe = StepResult(
                status="success" if text else "failed",
                file=text if text else None,
                error=err,
                retries_used=retries,
            )

            if not text:
                result.overall_status = "partial"
                result.error = f"下载+转码成功但识别失败: {err}"
                return result

        # 识别成功 → 落盘 transcript 文件（断点续跑校验依据）
        try:
            tp = transcript_path(sheet_name, stem)
            tp.write_text(result.transcribe.file, encoding="utf-8")
        except OSError as e:
            log.warning(f"[{stem}] 写入 transcript 失败: {e}")
    else:
        # 回退：transcribe 不在 steps 中（或无音频文件）时，从磁盘加载已落盘的 transcript
        _tp = transcript_path(sheet_name, stem)
        if _tp.exists():
            try:
                _cached = _tp.read_text(encoding="utf-8")
                ok, _ = validate_transcript_text(_cached)
                if ok:
                    result.transcribe = StepResult("success", file=_cached)
                    log.info(f"[{stem}] transcribe 未执行，从磁盘加载 transcript ({len(_cached)} 字符)")
            except OSError:
                pass

    # ── AI 分析（transcribe 之后执行）──
    if "analyze" in steps and result.transcribe.status == "success":
        ai_enabled = os.getenv("AI_ENABLED", "true").lower() == "true"
        if ai_enabled:
            if "analyze" in skip_steps:
                kw = result.analyze.file
                with _print_lock:
                    print(c("dim", f"  [{stem}] ♻️ 跳过 analyze，复用缓存关键词({len(kw)}字符)"), flush=True)
                result.analyze = StepResult("success", file=kw)
            else:
                txt = result.transcribe.file
                if txt:
                    try:
                        ai_start = time.monotonic()
                        kw, retries, err = step_analyze(txt, max_retries, retry_delay, analyze_timeout, result.stem)
                    except Exception as e:
                        kw, retries, err = None, max_retries, str(e)[:500]

                    # 校验：analyze 不允许半成功
                    ok, validate_err = validate_keywords_text(kw)
                    if not ok:
                        kw = None
                        err = err or validate_err
                        safe_remove(keywords_path(sheet_name, stem))

                    result.analyze = StepResult(
                        status="success" if kw else "failed",
                        file=kw,
                        error=err,
                        retries_used=retries,
                    )
                    if kw:
                        print(f"  [{result.stem}] {c('green', 'AI 分析完成')}（{fmt_elapsed(time.monotonic() - ai_start)}, {len(kw)} 字符）", flush=True)
                    else:
                        print(f"  [{result.stem}] {c('red', 'AI 分析失败')}：{err}", flush=True)
                else:
                    result.analyze = StepResult("skipped", error="识别文本为空")

            # analyze 成功 → 落盘 keywords 文件（断点续跑校验依据）
            if result.analyze.status == "success" and result.analyze.file:
                try:
                    kp = keywords_path(sheet_name, stem)
                    kp.write_text(result.analyze.file, encoding="utf-8")
                except OSError as e:
                    log.warning(f"[{stem}] 写入 keywords 失败: {e}")
        else:
            result.analyze = StepResult("skipped")
    elif "analyze" in steps and result.transcribe.status != "success":
        result.analyze = StepResult("skipped", error="transcribe 未成功，跳过 AI 分析")

    # ── 统一判定整体状态（和本地文件模式一致）──
    if result.transcode.status == "failed":
        result.overall_status = "failed"
    elif result.transcribe.status == "failed" and "transcribe" in steps:
        result.overall_status = "partial"
    elif result.analyze.status == "failed":
        result.overall_status = "partial"
    elif result.overall_status == "pending":
        result.overall_status = "success"

    return result


# ─────────────────────────────── 主控流程 ───────────────────────────────────

# ═══════════════════════════════════════════════════════════════════
# URL 直链流水线（--url 模式）— 将被注入到 process_videos.py
# ═══════════════════════════════════════════════════════════════════

def _run_url_task(opts):
    """执行 URL 直链的完整流水线。"""
    watch_url = opts["watch_url"]
    platform = opts["platform"]
    pkey = opts["pkey"]
    video_id = opts["video_id"]
    stem = opts["stem"]
    steps = opts["steps"]
    max_retries = opts["max_retries"]
    retry_delay = opts["retry_delay"]
    force = opts["force"]
    download_timeout = opts["download_timeout"]
    transcode_timeout = opts["transcode_timeout"]
    transcribe_timeout = opts["transcribe_timeout"]
    analyze_timeout = opts["analyze_timeout"]

    sheet_name = platform
    platform_field = PLATFORM_CONFIG[pkey].get("field", "") if pkey in PLATFORM_CONFIG else ""

    # 构建合成 row: 一个伪装成 Excel 行的 Series
    import pandas as pd
    synthetic_row = pd.Series(dtype=object)
    if platform_field:
        synthetic_row[platform_field] = video_id
    synthetic_row[COL_ID] = video_id
    synthetic_row[COL_TITLE] = video_id
    # 预缓存 stem（url-tasks, 0），供 stem_name 查找
    _STEM_CACHE[("url-tasks", 0)] = stem

    print(c("dim", "\n── 开始执行 ──\n"))

    result = process_one_task(
        synthetic_row,
        sheet_name,  # 使用平台名作为 sheet_name，按站点组织转码文件
        steps,
        max_retries,
        retry_delay,
        force,
        True,  # whisper_available（上游已检查）
        "",
        download_timeout,
        transcode_timeout,
        transcribe_timeout,
        analyze_timeout,
    )

    # ── 展示结果 ──
    print(c("dim", "\n── 结果 ──\n"))
    successes = []

    dl = result.download
    if dl and dl.file and Path(dl.file).exists():
        size_mb = Path(dl.file).stat().st_size / 1024 / 1024
        print(f"  📥 {c('green', '下载')}: {dl.file} ({size_mb:.1f} MB)")
        successes.append("download")
    elif dl and dl.status == "skipped":
        print(c("dim", "  📥 下载: 已跳过 (文件已存在)"))
        successes.append("download")
    elif dl:
        print(f"  📥 {c('red', '下载失败')}")
        _print_long(dl.error, indent="       ")

    tc = result.transcode
    if tc and tc.file and Path(tc.file).exists():
        size_mb = Path(tc.file).stat().st_size / 1024 / 1024
        print(f"  🎵 {c('green', '转码')}: {tc.file} ({size_mb:.1f} MB)")
        successes.append("transcode")
    elif tc and tc.status == "skipped":
        print(c("dim", "  🎵 转码: 已跳过 (文件已存在)"))
        successes.append("transcode")
    elif tc:
        print(f"  🎵 {c('red', '转码失败')}")
        _print_long(tc.error, indent="       ")

    tr = result.transcribe
    if tr and tr.file and isinstance(tr.file, str):
        print(f"  📝 {c('green', '识别')}: {len(tr.file)} 字符")
        successes.append("transcribe")
    elif tr and tr.status == "skipped":
        print(c("dim", "  📝 识别: 已跳过"))
        successes.append("transcribe")
    elif tr:
        print(f"  📝 {c('red', '识别失败')}")
        _print_long(tr.error, indent="       ")

    an = result.analyze
    if an and an.file and isinstance(an.file, str):
        print(f"  🤖 {c('green', 'AI分析')}: {len(an.file)} 字符")
        successes.append("analyze")
    elif an and an.status == "skipped":
        print(c("dim", "  🤖 AI分析: 已跳过"))
    elif an:
        print(f"  🤖 {c('red', 'AI分析失败')}")
        _print_long(an.error, indent="       ")

    # 保存文本结果
    transcribe_text = tr.file if (tr and isinstance(tr.file, str)) else ""
    analyze_text = an.file if (an and isinstance(an.file, str)) else ""

    if transcribe_text or analyze_text:
        out_dir = REPORTS_DIR / platform / "tasks"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f"{stem}.txt"
        lines = [
            f"URL: {watch_url}",
            f"平台: {platform}",
            f"视频ID: {video_id}",
            "", "=" * 60, "",
        ]
        if transcribe_text:
            lines.extend(["【语音识别内容】", "", transcribe_text, ""])
        if analyze_text:
            lines.extend(["【AI关键词分析】", "", analyze_text, ""])
        out_file.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n  📄 {c('cyan', '结果已保存至')}: {out_file}")

    print(c("bold", c("green", f"\n🎉 全部完成! ({len(successes)}/{len(steps)} 步成功)\n")))

    # 生成标准报告 JSON（与 Excel 模式格式一致）
    _generate_report_for_result(result, {"steps": steps, "max_retries": max_retries,
                                    "retry_delay": retry_delay, "concurrency": 1, "force": force},
                                   sheet_name=platform)

def _generate_report_for_result(result, config, sheet_name=None):
    """为单个 TaskResult 生成标准报告 JSON + 控制台摘要"""
    if result is None:
        return
    report_path = generate_report([result], config, sheet_name=sheet_name)
    print_report_summary([result])


def run(
    target_sheet: str | None,
    target_ids: list[str] | None,
    steps: list[str],
    max_retries: int,
    retry_delay: float,
    concurrency: int,
    force: bool,
    dry_run: bool,
    retry_failed: str | None,
    download_timeout: int = 1800,
    transcode_timeout: int = 1200,
    transcribe_timeout: int = 0,
    analyze_timeout: int = 300,
    offset: int = 0,
    limit: int = 0,
):
    """主执行流程"""
    # ── 重跑失败模式 ──
    if retry_failed:
        return run_from_report(retry_failed, steps, max_retries, retry_delay,
                               concurrency, force, dry_run,
                               download_timeout, transcode_timeout, transcribe_timeout,
                               analyze_timeout)

    # ── 构建任务列表 ──
    sheets = [target_sheet] if target_sheet else VIDEO_SHEETS
    tasks = []
    for sheet_name in sheets:
        df = pd.read_excel(str(EXCEL_FILE), sheet_name=sheet_name)
        if target_ids:
            mask = pd.Series([False] * len(df))
            # ID 列候选名：COL_ID、常见变体
            id_col_candidates = list(dict.fromkeys([COL_ID, 'id', 'ID', 'Id']))
            for _tid in target_ids:
                matched_any_col = False
                for col in id_col_candidates:
                    if col in df.columns:
                        try:
                            mask = mask | (df[col].apply(
                                lambda x: str(int(float(x))) if pd.notna(x) else ""
                            ) == str(_tid))
                            matched_any_col = True
                        except Exception:
                            pass
                if not matched_any_col and COL_TITLE in df.columns:
                    mask = mask | (df[COL_TITLE].astype(str) == str(_tid))
            df = df[mask]
            if df.empty:
                available_cols = list(df.columns)
                log.error(f"Sheet [{sheet_name}] 未找到匹配 --id 的行: {target_ids}")
                log.error(f"  可用列: {available_cols}")
                log.error(f"  ID 列候选: {id_col_candidates}  →  请确认 .env 中 COL_ID 是否和 Excel 列名一致")
                continue
        # 预计算 stems（同 sheet 内去重）
        precompute_stems(df, sheet_name)
        for _, row in df.iterrows():
            tasks.append((row, sheet_name))

    # ── 偏移/限量（全局，跨 sheet） ──
    if offset > 0 or limit > 0:
        start = offset
        end = start + limit if limit > 0 else None
        original_len = len(tasks)
        tasks = tasks[start:end]
        log.info(f"应用 offset={start}, limit={limit or '全部'} → 任务: {original_len} → {len(tasks)}")

    log.info(f"任务数量: {len(tasks)}，并发数: {concurrency}，最大重试: {max_retries}")

    # ── 工具/服务预检 ──
    if not _check_and_confirm_env(steps, dry_run, "是否继续执行？"):
        return

    # ── 干跑模式 ──
    if dry_run:
        print("\n" + "=" * 60)
        print(f"  干跑模式 - 任务清单 ({len(tasks)} 条)")
        print("=" * 60)
        env = check_environment(steps)  # 获取工具状态用于每步标记

        # ── 任务列表 ──
        print("\n  --- 任务步骤状态 ---")
        for i, (row, sheet_name) in enumerate(tasks):
            pkey, vid = get_video_id(row)
            stem = stem_name(row, sheet_name)
            url = build_url(pkey, vid) if pkey else "N/A"

            dl_path = (DOWNLOADS_DIR / sheet_name / stem).with_suffix(".mp4")
            tc_path = TRANSCODED_DIR / sheet_name / (stem + TRANSCODE_EXT)
            dl_exists = dl_path.exists()
            tc_exists = tc_path.exists()

            # 检查 Excel 列是否已填写
            content_val = row.get(COL_CONTENT)
            content_filled = pd.notna(content_val) and str(content_val).strip() != ""
            keywords_val = row.get(COL_KEYWORDS)
            keywords_filled = pd.notna(keywords_val) and str(keywords_val).strip() != ""

            print(f"\n  {i+1}. [{sheet_name}] {stem}")
            print(f"     platform={pkey}, url={url}")

            if not pkey:
                print(f"     ⚠️ 无可用视频 ID")
                continue

            # 每步状态
            if "download" in steps:
                if dl_exists:
                    status = "[跳过-已有文件]"
                elif not env["ytdlp"]:
                    status = "[不可用-yt-dlp]"
                else:
                    status = "[待执行]"
                print(f"      download : {status}")

            if "transcode" in steps:
                if tc_exists:
                    status = "[跳过-已有文件]"
                elif not env["ffmpeg"]:
                    status = "[不可用-ffmpeg]"
                elif not dl_exists:
                    status = "[等待-需先下载]"
                else:
                    status = "[待执行]"
                print(f"      transcode: {status}")

            if "transcribe" in steps:
                if content_filled:
                    status = f"[跳过-content已有{len(str(content_val))}字符]"
                elif not env["whisper"]:
                    status = "[不可用-whisper]"
                elif not tc_exists:
                    status = "[等待-需先转码]"
                else:
                    status = "[待执行]"
                print(f"      transcribe: {status}")

            if "analyze" in steps:
                if keywords_filled:
                    status = f"[跳过-keywords已有{len(str(keywords_val))}字符]"
                elif not env["ai"]:
                    status = "[不可用-AI未配置]"
                elif not content_filled and not tc_exists:
                    status = "[等待-需先识别]"
                else:
                    status = "[待执行]"
                print(f"      analyze  : {status}")

        return

    # ── 检测 whisper ──
    whisper_available = _check_whisper_available() if "transcribe" in steps else False
    if "transcribe" in steps and not whisper_available:
        if WHISPER_BACKEND == "local":
            backend_info = "本地 whisper CLI"
        elif WHISPER_BACKEND == "faster-whisper":
            backend_info = "faster-whisper (faster_whisper)"
        elif WHISPER_BACKEND == "funasr":
            backend_info = f"funasr/{FUNASR_MODE}"
        else:
            backend_info = WHISPER_SERVICE
        log.warning(f"⚠️ whisper 不可用 ({backend_info})，识别步骤将跳过")

    # ── 断点续跑：扫描 progress JSON，统计将跳过的任务/步骤数 ──
    if not force:
        skipped_tasks = 0
        resume_tasks = 0
        for (row, sn) in tasks:
            prior = load_task_progress(sn, stem_name(row, sn))
            if not prior:
                continue
            if prior.get("overall_status") == "success":
                skipped_tasks += 1
            else:
                resume_tasks += 1
        if skipped_tasks or resume_tasks:
            print()
            print(c("cyan", f"  ♻️ 断点续跑扫描: 完整跳过 {skipped_tasks} 条 / 部分续跑 {resume_tasks} 条 / 全量重跑 {len(tasks) - skipped_tasks - resume_tasks} 条"))
            print()

    # ── 并发执行 ──
    results: list[TaskResult] = []
    overall = OverallProgress(total=len(tasks))
    # 提前收集 content / keywords 用于 Excel 回写（save_task_progress 后会释放内存）
    content_updates: dict[tuple[str, str], str] = {}
    ai_updates: dict[tuple[str, str], str] = {}

    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as executor:
        future_map = {}
        for idx, (row, sheet_name) in enumerate(tasks):
            pos_label = f"[{idx + 1}/{len(tasks)}]"
            future_map[
                executor.submit(
                    process_one_task, row, sheet_name, steps,
                    max_retries, retry_delay, force, whisper_available,
                    pos_label,
                    download_timeout, transcode_timeout, transcribe_timeout, analyze_timeout,
                )
            ] = (row, sheet_name)

        for future in as_completed(future_map):
            try:
                result = future.result()
                results.append(result)
                overall.add_result(result.overall_status)
                # ── 收集 content / keywords 再保存进度（保存后会释放内存）──
                if result.transcribe.status == "success" and isinstance(result.transcribe.file, str):
                    content_updates[(result.sheet, result.id_val)] = result.transcribe.file
                if result.analyze.status == "success" and isinstance(result.analyze.file, str):
                    ai_updates[(result.sheet, result.id_val)] = result.analyze.file
                save_task_progress(result)
            except Exception as e:
                row, sheet_name = future_map[future]
                stem = stem_name(row, sheet_name)
                log.error(f"[{stem}] 任务执行异常: {e}\n{traceback.format_exc()}")
                tr = TaskResult(
                    sheet=sheet_name, id_val=row_key(row),
                    title=str(row.get(COL_TITLE, "")), stem=stem,
                    overall_status="failed", error=f"未捕获异常: {str(e)[:500]}",
                )
                results.append(tr)
                overall.add_result("failed")
                save_task_progress(tr)

            # 每次完成打印分隔线 + 总体进度
            with _print_lock:
                print()
                print(c("dim", "─" * 62))
                print(overall.summary_line(), flush=True)
                print()

    # ── 批量写回 Excel ──
    if "transcribe" in steps:
        write_all_contents_to_excel(results, ai_updates if ai_updates else None,
                                     content_updates if content_updates else None)

    # ── 生成报告 ──
    config = {
        "sheets": sheets,
        "target_id": target_id,
        "steps": steps,
        "max_retries": max_retries,
        "retry_delay": retry_delay,
        "concurrency": concurrency,
        "force": force,
    }
    report_path = generate_report(results, config)
    print_report_summary(results)

    log.info(f"全部完成！报告: {report_path}")


def run_from_report(
    report_path: str, steps: list[str],
    max_retries: int, retry_delay: float,
    concurrency: int, force: bool, dry_run: bool,
    download_timeout: int = 1800,
    transcode_timeout: int = 1200,
    transcribe_timeout: int = 0,
    analyze_timeout: int = 300,
):
    """从报告加载失败项，重新执行"""
    with open(report_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    failed_items = report.get("failed_items", [])
    if not failed_items:
        log.info("报告中没有失败项，无需重跑")
        return

    log.info(f"从报告加载 {len(failed_items)} 条失败项")

    # 按 sheet 分组，预计算 stems（处理去重）
    sheet_dfs: dict[str, pd.DataFrame] = {}
    tasks = []
    for item in failed_items:
        sheet_name = item["sheet"]
        key = str(item["id"])

        if sheet_name not in sheet_dfs:
            sheet_dfs[sheet_name] = pd.read_excel(str(EXCEL_FILE), sheet_name=sheet_name)

        df = sheet_dfs[sheet_name]
        # 匹配行
        mask = pd.Series([False] * len(df))
        if COL_ID in df.columns:
            try:
                mask = mask | (df[COL_ID].apply(
                    lambda x: str(int(float(x))) if pd.notna(x) else ""
                ) == key)
            except Exception:
                pass
        if COL_TITLE in df.columns:
            mask = mask | (df[COL_TITLE].astype(str) == key)

        matched = df[mask]
        if matched.empty:
            log.warning(f"[{sheet_name}] 找不到 {key}，跳过")
            continue
        tasks.append((matched.iloc[0], sheet_name))

    # 预计算 stems（按 sheet 去重）
    for sheet_name, sdf in sheet_dfs.items():
        precompute_stems(sdf, sheet_name)

    if not tasks:
        log.info("无有效失败项可重跑")
        return

    if dry_run:
        print(f"\n  干跑模式 - 重跑 {len(tasks)} 条失败项")
        for i, (row, sheet_name) in enumerate(tasks):
            pkey, vid = get_video_id(row)
            stem = stem_name(row, sheet_name)
            url = build_url(pkey, vid) if pkey else "N/A"
            print(f"  {i+1}. [{sheet_name}] {stem}  platform={pkey}  url={url}")
        return

    # ── 工具/服务预检 ──
    if not _check_and_confirm_env(steps, dry_run, "是否继续重跑？"):
        return

    whisper_available = _check_whisper_available() if "transcribe" in steps else False

    results: list[TaskResult] = []
    overall = OverallProgress(total=len(tasks))
    content_updates: dict[tuple[str, str], str] = {}
    ai_updates: dict[tuple[str, str], str] = {}

    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as executor:
        future_map = {}
        for idx, (row, sheet_name) in enumerate(tasks):
            pos_label = f"[{idx + 1}/{len(tasks)}]"
            future_map[
                executor.submit(
                    process_one_task, row, sheet_name, steps,
                    max_retries, retry_delay, force, whisper_available,
                    pos_label,
                    download_timeout, transcode_timeout, transcribe_timeout, analyze_timeout,
                )
            ] = (row, sheet_name)

        for future in as_completed(future_map):
            try:
                result = future.result()
                results.append(result)
                overall.add_result(result.overall_status)
                # ── 收集 content / keywords 再保存进度（保存后会释放内存）──
                if result.transcribe.status == "success" and isinstance(result.transcribe.file, str):
                    content_updates[(result.sheet, result.id_val)] = result.transcribe.file
                if result.analyze.status == "success" and isinstance(result.analyze.file, str):
                    ai_updates[(result.sheet, result.id_val)] = result.analyze.file
                save_task_progress(result)
            except Exception as e:
                row, sheet_name = future_map[future]
                stem = stem_name(row, sheet_name)
                log.error(f"[{stem}] 任务执行异常: {e}")
                tr = TaskResult(
                    sheet=sheet_name, id_val=row_key(row),
                    title=str(row.get(COL_TITLE, "")), stem=stem,
                    overall_status="failed", error=str(e)[:500],
                )
                results.append(tr)
                overall.add_result("failed")
                save_task_progress(tr)

            with _print_lock:
                print(f"\n{overall.summary_line()}\n", flush=True)

    if "transcribe" in steps:
        write_all_contents_to_excel(results, ai_updates if ai_updates else None,
                                     content_updates if content_updates else None)

    config = {
        "mode": "retry-failed",
        "source_report": report_path,
        "steps": steps,
        "max_retries": max_retries,
        "concurrency": concurrency,
        "force": force,
    }
    report_path_new = generate_report(results, config)
    print_report_summary(results)
    log.info(f"重跑完成！报告: {report_path_new}")
# ─────────────────────────────── 本地文件流水线（--input 模式）───────────────────────────────────────

def validate_input_file(file_path):
    """验证本地视频文件，检测可执行步骤"""
    result = {
        'valid': False, 'format': '', 'has_video': False, 'has_audio': False,
        'video_codec': '', 'audio_codec': '', 'duration': 0, 'width': 0, 'height': 0,
        'errors': [], 'feasible_steps': [],
    }

    # 1. 文件存在性
    path_obj = Path(file_path)
    if not path_obj.exists():
        result['errors'].append('文件不存在')
        return result

    if not path_obj.is_file():
        result['errors'].append('不是一个文件')
        return result

    if path_obj.stat().st_size == 0:
        result['errors'].append('文件大小为 0')
        return result

    # 2. ffprobe 分析
    if not shutil.which(FFPROBE):
        result['errors'].append(f'ffprobe 不可用 ({FFPROBE})')
        result['valid'] = True
        result['feasible_steps'] = ['transcode', 'transcribe', 'analyze']
        return result

    try:
        probe_cmd = [
            FFPROBE, '-v', 'error',
            '-show_entries', 'stream=codec_type,codec_name,width,height',
            '-show_entries', 'format=format_name,duration',
            '-of', 'json',
            str(path_obj)
        ]
        probe_raw = subprocess.check_output(probe_cmd, timeout=30, text=True)
        info = json.loads(probe_raw)

        # 提取 format 信息
        if 'format' in info:
            result['format'] = (info['format'].get('format_name', '') or '').split(',')[0]
            result['duration'] = float(info['format'].get('duration', 0) or 0)

        # 提取 stream 信息
        if 'streams' in info:
            for s in info['streams']:
                if s.get('codec_type') == 'video':
                    result['has_video'] = True
                    result['video_codec'] = s.get('codec_name', '')
                    result['width'] = s.get('width', 0)
                    result['height'] = s.get('height', 0)
                if s.get('codec_type') == 'audio':
                    result['has_audio'] = True
                    result['audio_codec'] = s.get('codec_name', '')

        result['valid'] = True

        # 3. 判断可执行步骤
        if result['has_video']:
            result['feasible_steps'].append('transcode')
        if result['has_audio']:
            result['feasible_steps'].extend(['transcribe', 'analyze'])

        # 无视频无音频 → 所有步骤不可行
        if not result['has_video'] and not result['has_audio']:
            result['feasible_steps'] = []
            result['errors'].append('文件不包含视频或音频流，无法处理')
        elif result['has_video'] and not result['has_audio']:
            result['errors'].append('文件不含音频轨道，将跳过语音识别和 AI 分析')

    except Exception as e:
        result['errors'].append(f'ffprobe 解析失败: {str(e)[:200]}')
        result['valid'] = True

    return result


def run_input_task(input_path, sheet_name, steps, max_retries, retry_delay, force,
                   transcode_timeout, transcribe_timeout, analyze_timeout, custom_name=None,
                   whisper_available=True):
    """--input 模式的独立流水线"""
    stem = safe_filename(custom_name) if custom_name else os.path.splitext(os.path.basename(str(input_path)))[0]

    # ── 解决 stem 重名 ──
    used_stem = stem
    tc_dir = TRANSCODED_DIR / sheet_name
    tc_dir.mkdir(parents=True, exist_ok=True)
    counter = 1
    if 'transcode' in steps and not force:
        test_path = tc_dir / (used_stem + TRANSCODE_EXT)
        while test_path.exists():
            used_stem = f"{stem}_{counter}"
            test_path = tc_dir / (used_stem + TRANSCODE_EXT)
            counter += 1
    if used_stem != stem:
        print(c("yellow", f"  ⚠️  stem '{stem}' 已存在 → 使用 '{used_stem}'"))

    result = TaskResult(
        sheet=sheet_name, id_val='-', title=input_path.name, stem=used_stem,
        platform='local', video_url=None,
        overall_status='pending', download=StepResult('skipped'),
    )

    tc_file = None

    # ── 转码 ──
    if 'transcode' in steps:
        try:
            tc_file, retries, err = step_transcode(
                input_path, sheet_name, max_retries, retry_delay, force, transcode_timeout,
                out_stem=used_stem,
            )
        except Exception as e:
            tc_file, retries, err = None, max_retries, str(e)[:500]

        result.transcode = StepResult(
            status='success' if tc_file else 'failed',
            file=str(tc_file) if tc_file else None,
            error=err,
            retries_used=retries,
        )

        if not tc_file:
            result.overall_status = 'failed'
            result.error = f'转码失败: {err}'
            return result
    else:
        # 不转码：如果有 transcribe 步骤，优先使用已有转码文件
        if 'transcribe' in steps:
            expected_tc = (Path(TRANSCODED_DIR) / sheet_name / (stem + TRANSCODE_EXT)).resolve()
            if expected_tc.exists():
                tc_file = expected_tc
                result.transcode = StepResult('success', file=str(tc_file))
            else:
                tc_file = input_path
                result.transcode = StepResult('warning', file=str(tc_file),
                                              error='未找到转码文件，将使用原始文件（识别可能失败）')
        else:
            tc_file = input_path
            result.transcode = StepResult('success', file=str(tc_file))

    # ── 识别 ──
    if 'transcribe' in steps and tc_file:
        if not whisper_available:
            result.transcribe = StepResult('failed', error=f'whisper 服务不可达 ({WHISPER_SERVICE})')
            result.overall_status = 'failed'
            result.error = '转码成功但 whisper 服务不可达'
            return result

        try:
            text_file, retries, err = step_transcribe(
                tc_file, max_retries, retry_delay, transcribe_timeout
            )
        except Exception as e:
            text_file, retries, err = None, max_retries, str(e)[:500]

        result.transcribe = StepResult(
            status='success' if text_file else 'failed',
            file=str(text_file) if text_file else None,
            error=err,
            retries_used=retries,
        )

        if not text_file:
            result.overall_status = 'partial'
            result.error = f'转码成功但识别失败: {err}'
            return result
    else:
        # 回退：transcribe 不在 steps 中时，从磁盘加载已落盘的 transcript
        _tp = transcript_path(sheet_name, result.stem)
        if _tp.exists():
            try:
                _cached = _tp.read_text(encoding="utf-8")
                ok, _ = validate_transcript_text(_cached)
                if ok:
                    result.transcribe = StepResult("success", file=_cached)
                    log.info(f"[{result.stem}] transcribe 未执行，从磁盘加载 transcript ({len(_cached)} 字符)")
            except OSError:
                pass

    # ── AI 分析 ──
    if 'analyze' in steps:
        if result.transcribe.status != 'success':
            result.analyze = StepResult('skipped', error='无可用 transcript（transcribe 未执行且磁盘无缓存）')
        else:
            try:
                ai_start = time.monotonic()
                analyze_ok, retries, err = step_analyze(
                    result.transcribe.file, max_retries, retry_delay, analyze_timeout, result.stem
                )
                if analyze_ok:
                    print(f"  [{result.stem}] {c('green', 'AI 分析完成')}（{fmt_elapsed(time.monotonic() - ai_start)}, {len(analyze_ok)} 字符）", flush=True)
            except Exception as e:
                analyze_ok, retries, err = False, max_retries, str(e)[:500]

            result.analyze = StepResult(
                status='success' if analyze_ok else 'failed',
                file=result.transcribe.file if analyze_ok else None,
                error=err,
                retries_used=retries,
            )

            if not analyze_ok:
                result.overall_status = 'partial'
                result.error = f'转码+识别成功但 AI 分析失败: {err}'
                return result
    else:
        result.analyze = StepResult('skipped')

    result.overall_status = 'success'

    # ── 保存文本结果 ──
    tr = result.transcribe
    an = result.analyze
    transcribe_text = tr.file if (tr and isinstance(tr.file, str)) else ""
    analyze_text = an.file if (an and isinstance(an.file, str)) else ""

    if transcribe_text or analyze_text:
        out_dir = REPORTS_DIR / sheet_name / "tasks"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f"{used_stem}.txt"
        lines = [
            f"文件: {input_path}",
            f"平台: local",
            "", "=" * 60, "",
        ]
        if transcribe_text:
            lines.extend(["【语音识别内容】", "", transcribe_text, ""])
        if analyze_text:
            lines.extend(["【AI关键词分析】", "", analyze_text, ""])
        out_file.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n  📄 {c('cyan', '结果已保存至')}: {out_file}")

    return result


# ─────────────────────────────── 纯文本 AI 分析（--content 模式） ─────────────────

def run_content_task(content, name, steps, max_retries, retry_delay, analyze_timeout,
                     force=False, dry_run=False):
    """--content 模式：纯文本 → AI 关键词提取"""
    content_path = Path(content).resolve()
    content_text = ""
    from_file = False

    if content_path.exists() and content_path.is_file():
        content_text = content_path.read_text(encoding="utf-8").strip()
        from_file = True
        print(f"  {c('dim', '从文件读取:')} {content_path} ({len(content_text)} 字符)")
    else:
        content_text = content

    if not content_text or not content_text.strip():
        print(c("red", "错误: --content 文本内容为空"), file=sys.stderr)
        sys.exit(1)

    # 确定输出名称
    if name:
        stem = safe_filename(name)
    elif from_file:
        stem = safe_filename(content_path.stem)
    else:
        stem = safe_filename(content_text.replace("\n", " ").replace("\r", "")[:32].strip())

    if not steps:
        steps = ["analyze"]

    print(c("dim", "\n── 开始执行 (内容分析) ──\n"))
    print(f"  输出名称:  {c('cyan', stem)}")
    print(f"  内容长度:  {c('cyan', str(len(content_text)) + ' 字符')}")
    print(f"  执行步骤:  {c('cyan', ' → '.join(steps))}")

    if dry_run:
        print("")
        sys.exit(0)

    sheet_name = "content"

    result = TaskResult(
        sheet=sheet_name,
        id_val="-",
        title=content[:50] if not from_file else content_path.name,
        stem=stem,
        platform="local",
        video_url=None,
        overall_status="pending",
        download=StepResult("skipped"),
        transcode=StepResult("skipped"),
        transcribe=StepResult("success", file=content_text),
    )

    # ── AI 分析 ──
    if "analyze" in steps:
        ai_enabled = os.getenv("AI_ENABLED", "true").lower() == "true"
        if not ai_enabled:
            result.analyze = StepResult("skipped")
            print(f"  [{stem}] {c('yellow', 'AI 分析已禁用 (AI_ENABLED=false)')}", flush=True)
        else:
            print(f"  [{stem}] {c('cyan', '开始 AI 分析...')}", flush=True)
            try:
                ai_start = time.monotonic()
                kw, retries, err = step_analyze(
                    content_text, max_retries, retry_delay, analyze_timeout, stem
                )
                if kw:
                    print(f"  [{stem}] {c('green', 'AI 分析完成')}（{fmt_elapsed(time.monotonic() - ai_start)}, {len(kw)} 字符）", flush=True)
                    result.analyze = StepResult("success", file=kw, retries_used=retries)
                else:
                    print(f"  [{stem}] {c('red', 'AI 分析失败')}: {err}", flush=True)
                    result.analyze = StepResult("failed", error=err, retries_used=retries)
            except Exception as e:
                result.analyze = StepResult("failed", error=str(e)[:500], retries_used=max_retries)
                print(f"  [{stem}] {c('red', 'AI 分析异常')}: {str(e)[:200]}", flush=True)
    else:
        result.analyze = StepResult("skipped")

    result.overall_status = "success" if (result.analyze and result.analyze.status == "success") else "partial"

    # ── 保存文本结果 ──
    an = result.analyze
    analyze_text = an.file if (an and an.status == "success" and isinstance(an.file, str)) else ""
    if content_text or analyze_text:
        out_dir = REPORTS_DIR / sheet_name / "tasks"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f"{stem}.txt"
        lines = [
            f"来源: --content",
            f"输出名称: {stem}",
            f"内容长度: {len(content_text)} 字符",
            "", "=" * 60, "",
            "【原始内容】", "", content_text, "",
        ]
        if analyze_text:
            lines.extend(["【AI 分析关键词】", "", analyze_text, ""])
        out_file.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n  {c('cyan', '报告已保存:')} {out_file}")

    # ── 总结 ──
    print("")
    success_parts = [label for label, cond in [
        ("analyze", analyze_text),
    ] if cond]
    failed_parts = [s for s in steps if s != "download" and s not in ["analyze"] and not success_parts]
    if not failed_parts and analyze_text:
        print(c("green", "✅ 全部步骤执行成功"))
    else:
        print(c("yellow", f"⚠️  部分步骤未成功"))

    return result


# ─────────────────────────────── 入口 ───────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="视频下载/转码/识别流程（支持并发、重试、报告）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python process_videos.py --concurrency 3 --retry 3
  python process_videos.py --sheet "YouTube视频" --id 2143
  python process_videos.py --retry-failed output/reports/YouTube视频/report_20260610_141800.json
  python process_videos.py --dry-run
        """,
    )
    parser.add_argument('--version', action='version', version=f'%(prog)s {__version__}')
    parser.add_argument("--sheet", help="指定 sheet 名称（默认全部视频 sheet）")
    parser.add_argument("--id", dest="vid_ids", action="append", default=[],
                        help="指定 extra.id 或 title，可多次指定或逗号分隔（如 --id 1,2,3 或 --id 1 --id 2）")
    parser.add_argument("--offset", type=int, default=0, help="跳过前 N 条任务（从 0 开始），默认 0")
    parser.add_argument("--limit", type=int, default=0, help="最多处理 N 条任务，默认 0 表示无限制")
    parser.add_argument(
        "--step", choices=["download", "transcode", "transcribe", "analyze"],
        action="append",
        help="指定执行步骤（可多次指定，如 --step transcode --step transcribe）",
    )
    parser.add_argument("--force", action="store_true", help="强制重新执行，忽略已有文件")
    parser.add_argument(
        "--concurrency", type=int, default=1,
        help="并发数，默认 1（仅配置 >1 时多线程并行，建议 2~3）",
    )
    parser.add_argument(
        "--retry", type=int, default=0,
        help="每个步骤失败后最大重试次数，默认 0",
    )
    parser.add_argument(
        "--retry-delay", type=float, default=5.0,
        help="重试间隔基数（秒），指数退避，默认 5s",
    )
    parser.add_argument(
        "--download-timeout", type=int, default=1800,
        help="单个下载任务的最长执行时间（秒），默认 1800s（30 分钟），设为 0 则不限制",
    )
    parser.add_argument(
        "--transcode-timeout", type=int, default=1200,
        help="单个转码任务的最长执行时间（秒），默认 1200s（20 分钟），设为 0 则不限制",
    )
    parser.add_argument(
        "--transcribe-timeout", type=int, default=0,
        help="单个识别任务的最长执行时间（秒），默认 0（不限制超时）",
    )
    parser.add_argument(
        "--analyze-timeout", type=int, default=300,
        help="单个 AI 分析任务的最长执行时间（秒），默认 300s（5 分钟），设为 0 则不限制",
    )
    parser.add_argument("--dry-run", action="store_true", help="干跑模式，仅列出任务不执行")
    parser.add_argument(
        "--retry-failed",
        help="从指定报告 JSON 重跑失败项（output/reports/{sheet}/report_xxx.json）",
    )
    parser.add_argument("--init", action="store_true", help="复制 .env.example 到当前目录并重命名为 .env")
    parser.add_argument(
        "--file",
        help="指定 Excel 文件路径（优先级高于 EXCEL_FILE 环境变量）",
    )
    parser.add_argument(
        "--env-file",
        help="指定要加载的 .env 文件路径（默认: 当前目录 .env）",
    )
    parser.add_argument(
        "--output",
        help="指定输出根目录（覆盖 OUTPUT_DIR 环境变量；子目录 downloads/transcoded/transcripts/keywords/reports/progress/logs 自动创建）",
    )
    parser.add_argument(
        "--url",
        help="直接指定视频下载链接（跳过 Excel），支持标准链接和内嵌链接",
    )
    parser.add_argument(
        "--name",
        help="指定下载文件名，不含扩展名（与 --url / --input 配合使用）",
    )
    parser.add_argument(
        "--input",
        help="指定本地视频文件路径（跳过下载，直接转码→识别→分析）",
    )
    parser.add_argument(
        "--content",
        help="直接提供文本内容（文件路径或内联文本），跳过下载/转码/识别，直接做 AI 分析",
    )
    parser.add_argument(
        "--whisper-initial-prompt",
        help="Whisper 初始提示词（文本或文件路径，CLI 优先级最高）",
    )
    parser.add_argument(
        "--ai-prompt",
        help="AI 分析提示词模板（文本或文件路径，CLI 优先级最高）",
    )
    parser.add_argument(
        "--whisper-extra-args",
        help='Whisper 额外参数（shell 字符串，如 "--beam_size 5 --best_of 5"，最高优先级且自动去重）',
    )
    args = parser.parse_args()

    # Windows 中文环境下强制 UTF-8，修复 yt-dlp / httpx / requests latin-1 编码错误
    if sys.platform == "win32":
        try:
            import ctypes
            k32 = ctypes.windll.kernel32
            k32.SetConsoleCP(65001)
            k32.SetConsoleOutputCP(65001)
        except Exception:
            pass
        os.environ["PYTHONUTF8"] = "1"
        os.environ["PYTHONIOENCODING"] = "utf-8"

    # ── CLI 覆盖：提示词文件/文本归一化 + whisper extra args ──
    apply_cli_overrides(args)

    # ── init 模式 ──
    if args.init:
        src = SCRIPT_DIR / ".env.example"
        if not src.exists():
            print(f"错误: 找不到 {src}", file=sys.stderr)
            sys.exit(1)
        import shutil as _shutil
        import questionary as _qy
        dest = Path.cwd() / ".env"
        if dest.exists():
            print(f"\n⚠️  目标文件已存在: {dest}")
            choice = _qy.select(
                "如何处理冲突?",
                choices=[
                    _qy.Choice("覆盖 (overwrite)", "overwrite"),
                    _qy.Choice("保留现有 (keep existing)", "keep"),
                    _qy.Choice("自定义文件名 (custom name)", "custom"),
                ],
            ).unsafe_ask()
            if choice == "overwrite":
                _shutil.copy2(str(src), str(dest))
                print(f"✅ .env 已覆盖: {dest}")
            elif choice == "custom":
                custom_name = _qy.text(
                    "请输入新文件名",
                    default=".env.prod",
                    validate=lambda v: bool(v.strip()),
                ).unsafe_ask()
                if not custom_name:
                    print("未输入文件名，已取消。")
                    sys.exit(0)
                dest = Path.cwd() / safe_filename(custom_name)
                if dest.exists():
                    print(f'⚠️  文件 "{custom_name}" 也已存在，保留现有文件。')
                else:
                    _shutil.copy2(str(src), str(dest))
                    print(f"✅ .env 已创建为: {dest}")
            else:
                print("保留现有 .env 文件，未做修改。")
        else:
            _shutil.copy2(str(src), str(dest))
            print(f"✅ .env 已从 .env.example 创建: {dest}")
        sys.exit(0)

    # ── file 覆盖 ──
    if args.file:
        EXCEL_FILE = Path(args.file).resolve()
        log.info(f"Excel 文件覆盖为: {EXCEL_FILE}")

    # ── output 覆盖（CLI > env > 默认 "output"）──
    if args.output:
        _new_root = Path(args.output).resolve() if Path(args.output).is_absolute() else (BASE_DIR / args.output).resolve()
        if _new_root != OUTPUT_DIR:
            _apply_output_dir(_new_root, log_func=log.info)

    steps = args.step if args.step else ["download", "transcode", "transcribe", "analyze"]

    # ── --input 模式：移除 download 步骤 ──
    if args.input and not args.step:
        steps = ["transcode", "transcribe", "analyze"]

    # ── --url 模式：直接处理单个视频链接 ──
    if args.url:
        parsed = parse_url(args.url)
        if not parsed:
            print(c("red", f"\n❌ 无法识别的 URL: {args.url}"))
            print(c("dim", "支持的平台: YouTube, B站, 腾讯视频, 优酷"))
            print(c("dim", "URL 格式示例:"))
            print(c("dim", "  https://www.bilibili.com/video/BV1xxxyyyzzz"))
            print(c("dim", "  https://www.youtube.com/watch?v=xxxxxxxxxxx"))
            print(c("dim", "  https://v.qq.com/x/page/x0000xxxxx.html"))
            print(c("dim", "  https://v.youku.com/v_show/id_XXXXXXX.html"))
            sys.exit(1)

        platform = parsed["platform"]
        video_id = parsed["video_id"]
        watch_url = parsed["watch_url"]
        pkey = parsed["pkey"]

        print(c("dim", "\n── URL 任务 ──"))
        print(f"  平台: {c('cyan', platform)}")
        print(f"  视频ID: {c('cyan', video_id)}")
        print(f"  链接: {c('cyan', watch_url)}")

        # dry-run 模式
        if args.dry_run:
            print(c("dim", "\n── 开始执行 (dry-run) ──\n"))
            print(f"  将执行步骤: {c('cyan', ' → '.join(steps))}")
            print(f"  输出名称: {c('cyan', args.name if args.name else video_id)}")
            sys.exit(0)

        # 构建文件路径: output/downloads/<platform>/<name>.mp4
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
        dl_dir = DOWNLOADS_DIR / platform
        dl_dir.mkdir(parents=True, exist_ok=True)
        file_name = safe_filename(args.name if args.name else video_id)
        proposed_path = dl_dir / f"{file_name}.mp4"

        # 冲突处理（--force 时直接覆盖）
        if args.force:
            final_path = proposed_path
            final_stem = file_name
        else:
            conflict = resolve_url_conflict(proposed_path)
            if conflict["action"] == "skip":
                print("\n⏭️  已跳过\n")
                sys.exit(0)
            final_path = conflict["path"]
            final_stem = final_path.stem

        print(f"  文件: {final_path}")

        # 检查 whisper 可用性
        whisper_available = True
        if "transcribe" in steps:
            whisper_available = _check_whisper_available()
            if not whisper_available:
                if WHISPER_BACKEND == "local":
                    backend = "local CLI"
                elif WHISPER_BACKEND == "faster-whisper":
                    backend = "faster-whisper (faster_whisper)"
                elif WHISPER_BACKEND == "funasr":
                    backend = f"funasr/{FUNASR_MODE}"
                else:
                    backend = WHISPER_SERVICE
                log.warning(f"⚠️ whisper not available ({backend}), transcribe step will fail")

        # 执行流水线
        _run_url_task({
            "watch_url": watch_url,
            "platform": platform,
            "pkey": pkey,
            "video_id": video_id,
            "stem": final_stem,
            "dl_dir": dl_dir,
            "steps": steps,
            "max_retries": args.retry,
            "retry_delay": args.retry_delay,
            "force": args.force,
            "download_timeout": args.download_timeout,
            "transcode_timeout": args.transcode_timeout,
            "transcribe_timeout": args.transcribe_timeout,
            "analyze_timeout": args.analyze_timeout,
        })

        sys.exit(0)

    # ── --input 模式：直接处理本地视频文件 ──
    if args.input:
        input_path = Path(args.input).resolve()

        print(c("dim", "\n── 文件校验 ──"))
        print(f"  文件: {c('cyan', str(input_path))}")

        file_info = validate_input_file(input_path)

        if not file_info['valid']:
            print(c("red", f"\n❌ 无法处理该文件:"))
            for err in file_info['errors']:
                print(c("red", f"   - {err}"))
            sys.exit(1)

        print(f"  格式: {c('cyan', file_info['format'] or 'unknown')}")
        if file_info['has_video']:
            video_info = f"{file_info['video_codec']} {file_info['width']}x{file_info['height']}"
            print(f"  视频: {c('cyan', video_info)}")
        if file_info['has_audio']:
            print(f"  音频: {c('cyan', file_info['audio_codec'])}")
        dur_label = f"{int(file_info['duration'] // 60)}:{int(file_info['duration'] % 60):02d}"
        print(f"  时长: {c('cyan', dur_label)} ({file_info['duration']:.1f}s)")

        if file_info['errors']:
            print(c("yellow", f"\n⚠️  警告:"))
            for err in file_info['errors']:
                print(c("yellow", f"   - {err}"))

        print(f"\n  {c('green', '可执行步骤')}: {c('cyan', ' → '.join(file_info['feasible_steps']))}")

        # 检查请求的步骤是否都可行
        for step in steps:
            if step not in file_info['feasible_steps']:
                print(c("red", f"\n❌ 错误: 文件不支持 '{step}' 步骤"))
                print(c("dim", f"   支持的步骤: {', '.join(file_info['feasible_steps'])}"))
                sys.exit(1)

        # 确定 sheet 名称（用于输出目录）
        sheet_name = args.sheet if args.sheet else "local"

        # 检查 whisper 可用性
        whisper_available = True
        if "transcribe" in steps:
            whisper_available = _check_whisper_available()
            if not whisper_available:
                if WHISPER_BACKEND == "local":
                    backend = "local CLI"
                elif WHISPER_BACKEND == "faster-whisper":
                    backend = "faster-whisper (faster_whisper)"
                elif WHISPER_BACKEND == "funasr":
                    backend = f"funasr/{FUNASR_MODE}"
                else:
                    backend = WHISPER_SERVICE
                log.warning(f"⚠️ whisper not available ({backend}), transcribe step will fail")

        # dry-run 模式
        if args.dry_run:
            print(c("dim", f"\n── 开始执行 (dry-run) ──\n"))
            print(f"  [本地文件] 将执行步骤: {c('cyan', ' → '.join(steps))}")
            print(f"  输入文件: {c('cyan', str(input_path))}")
            if args.name:
                print(f"  输出名称: {c('cyan', args.name)}")
            sys.exit(0)

        # 执行流水线
        result = run_input_task(
            input_path=input_path,
            sheet_name=sheet_name,
            steps=steps,
            max_retries=args.retry,
            retry_delay=args.retry_delay,
            force=args.force,
            transcode_timeout=args.transcode_timeout,
            transcribe_timeout=args.transcribe_timeout,
            analyze_timeout=args.analyze_timeout,
            custom_name=args.name,
            whisper_available=whisper_available,
        )

        # 生成标准报告 JSON（与 Excel 模式格式一致）
        config = {
            "steps": steps,
            "max_retries": args.retry,
            "retry_delay": args.retry_delay,
            "concurrency": 1,
            "force": args.force,
        }
        report_path = generate_report([result], config, sheet_name=sheet_name)
        print_report_summary([result])
        print(f"  整体状态: {c('green' if result.overall_status == 'success' else 'red', result.overall_status)}")
        if result.download:
            print(f"  {c('dim', '下载')}: {result.download.status}")
        if result.transcode:
            print(f"  {c('dim', '转码')}: {result.transcode.status}")
            if result.transcode.file:
                print(f"    {c('dim', '输出')}: {result.transcode.file}")
        if result.transcribe:
            print(f"  {c('dim', '识别')}: {result.transcribe.status}")
            if result.transcribe.file:
                print(f"    {c('dim', '输出')}: {result.transcribe.file}")
        if result.analyze:
            print(f"  {c('dim', '分析')}: {result.analyze.status}")

        if result.error:
            print(f"\n  错误: {result.error}")

        sys.exit(0)

    # ── --content 模式：纯文本 AI 分析 ──
    if args.content:
        result = run_content_task(
            content=args.content,
            name=args.name,
            steps=steps,
            max_retries=args.retry,
            retry_delay=args.retry_delay,
            analyze_timeout=args.analyze_timeout,
            force=args.force,
            dry_run=args.dry_run,
        )
        sys.exit(0)

    run(
        target_sheet=args.sheet,
        target_ids=[s for v in (args.vid_ids or []) for s in str(v).split(",")] or None,
        steps=steps,
        offset=args.offset,
        limit=args.limit,
        max_retries=args.retry,
        retry_delay=args.retry_delay,
        concurrency=args.concurrency,
        force=args.force,
        dry_run=args.dry_run,
        retry_failed=args.retry_failed,
        download_timeout=args.download_timeout,
        transcode_timeout=args.transcode_timeout,
        transcribe_timeout=args.transcribe_timeout,
        analyze_timeout=args.analyze_timeout,
    )
